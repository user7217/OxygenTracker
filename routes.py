"""
Varasai Oxygen Cylinder Tracker - Flask Routes and Application Logic

Flask routes and application logic for the cylinder management system.

Features:
- Complete CRUD operations for customers and cylinders
- Advanced pagination system for large datasets (5000+ cylinders)
- Role-based access control (Admin, User, Viewer)
- Search and filtering capabilities
- Bulk cylinder operations and rental management
- Data import from MS Access databases
- CSV and PDF export functionality
- Transaction management for customer-cylinder relationships
- Responsive web interface with mobile optimization

Route Categories:
- Authentication: Login, logout, user management
- Dashboard: Main interface and metrics
- Customers: Customer management with Access-compatible fields
- Cylinders: Cylinder inventory with rental tracking and pagination
- Bulk Operations: Multi-cylinder rental/return operations
- Import/Export: Data migration and reporting
- Search: Global search across customers and cylinders
- Admin: User management and system administration

Author: Development Team
Date: July 2025
Version: 2.0
"""

from flask import render_template, request, redirect, url_for, flash, jsonify, session, Response
import csv
import io
import os
import json
import shutil
import threading
import time
from datetime import datetime, timedelta
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from app import app
from db_service import CustomerService, CylinderService, RentalHistoryService
from auth_models import UserManager
from functools import wraps
import os
import tempfile
import json

# Try to import Access functionality with graceful degradation
# MS Access import is optional - system works without it

# JSON Import functionality
from json_importer import JSONImporter
try:
    from data_importer import DataImporter
    ACCESS_AVAILABLE = True
except ImportError as e:
    ACCESS_AVAILABLE = False
    import logging
    logging.warning(f"MS Access functionality not available: {e}")

# Try to import Email functionality with graceful degradation
# Email service is optional - system works without it
try:
    from email_service import EmailService
    email_service = EmailService()
    EMAIL_AVAILABLE = True
except ImportError as e:
    EMAIL_AVAILABLE = False
    email_service = None
    import logging
    logging.warning(f"Email functionality not available: {e}")

# Initialize user manager for authentication and authorization
user_manager = UserManager()

def login_required(f):
    """
    Decorator to require user authentication for routes
    
    This decorator ensures that only authenticated users can access protected routes.
    Redirects unauthenticated users to the login page with appropriate flash message.
    
    Args:
        f (function): The route function to protect
        
    Returns:
        function: Wrapped function with authentication check
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    """
    Decorator to require admin role for routes
    
    This decorator ensures that only users with admin role can access admin-only routes.
    Provides the highest level of access control for sensitive operations like user
    management, data exports, and system administration.
    
    Args:
        f (function): The route function to protect
        
    Returns:
        function: Wrapped function with admin role check
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page', 'error')
            return redirect(url_for('login'))
        
        user = user_manager.get_user_by_id(session['user_id'])
        if not user or user.get('role') != 'admin':
            flash('Admin access required', 'error')
            return redirect(url_for('index'))
        
        return f(*args, **kwargs)
    return decorated_function

def user_or_admin_required(f):
    """
    Decorator to require user or admin role (excludes viewers)
    
    This decorator allows access to users with 'user' or 'admin' roles while
    excluding viewers from operational functions like cylinder rental/return,
    bulk operations, and data modifications.
    
    Args:
        f (function): The route function to protect
        
    Returns:
        function: Wrapped function with user/admin role check
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page', 'error')
            return redirect(url_for('login'))
        
        user = user_manager.get_user_by_id(session['user_id'])
        if not user or user.get('role') not in ['admin', 'user']:
            flash('Insufficient permissions', 'error')
            return redirect(url_for('index'))
        
        return f(*args, **kwargs)
    return decorated_function

def admin_or_user_can_edit(f):
    """
    Decorator for routes that require admin access for data modification
    
    This decorator restricts access to admin-only operations like adding,
    editing, or deleting customers and cylinders. Users can only perform
    rental/return operations but cannot modify core data.
    
    Args:
        f (function): The route function to protect
        
    Returns:
        function: Wrapped function with admin-only access check
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page', 'error')
            return redirect(url_for('login'))
        
        user = user_manager.get_user_by_id(session['user_id'])
        if not user or user.get('role') != 'admin':
            flash('Only administrators can add/edit/delete customers and cylinders', 'error')
            return redirect(url_for('index'))
        
        return f(*args, **kwargs)
    return decorated_function

# Using PostgreSQL services instead of model instances

# ============================================================================
# AUTHENTICATION ROUTES
# ============================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    User authentication and session management
    
    Handles user login with username/password authentication. Creates secure
    session with user ID, username, and role information. Supports redirect
    to requested page after successful login.
    
    GET: Display login form
    POST: Process login credentials and create session
    
    Returns:
        GET: Login template
        POST: Redirect to dashboard or requested page on success, login form on failure
    """
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        
        # Validate required fields
        if not username or not password:
            flash('Please enter both username and password', 'error')
            return render_template('login.html')
        
        # Authenticate user credentials
        user = user_manager.authenticate(username, password)
        if user:
            # Create secure session
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user.get('role', 'user')
            
            flash(f'Welcome back, {user["username"]}!', 'success')
            
            # Handle redirect to originally requested page
            next_page = request.args.get('next')
            if next_page:
                return redirect(next_page)
            return redirect(url_for('index'))
        else:
            flash('Invalid username or password', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """
    User logout and session cleanup
    
    Clears all session data and redirects to login page with farewell message.
    Ensures complete session cleanup for security.
    
    Returns:
        Redirect to login page with logout message
    """
    username = session.get('username', 'User')
    session.clear()
    flash(f'Goodbye, {username}!', 'info')
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
@admin_required
def register():
    """User registration (admin only)"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        role = request.form.get('role', 'viewer').strip()
        
        # Validation
        if not all([username, email, password, confirm_password]):
            flash('All fields are required', 'error')
            return render_template('register.html')
        
        if password != confirm_password:
            flash('Passwords do not match', 'error')
            return render_template('register.html')
        
        if len(password) < 6:
            flash('Password must be at least 6 characters long', 'error')
            return render_template('register.html')
        
        if role not in ['admin', 'user', 'viewer']:
            flash('Invalid role selected', 'error')
            return render_template('register.html')
        
        # Check if user already exists
        if user_manager.get_user_by_username(username):
            flash('Username already exists', 'error')
            return render_template('register.html')
        
        # Check if email already exists
        users = user_manager.get_all_users()
        if any(u.get('email') == email for u in users):
            flash('Email already registered', 'error')
            return render_template('register.html')
        
        try:
            new_user = user_manager.create_user(username, email, password, role)
            flash(f'User {username} created successfully with role: {role}', 'success')
            return redirect(url_for('users'))
        except Exception as e:
            flash(f'Error creating user: {str(e)}', 'error')
    
    return render_template('register.html')

@app.route('/users')
@admin_required
def users():
    """List all users (admin only)"""
    all_users = user_manager.get_all_users()
    return render_template('users.html', users=all_users)

# ============================================================================
# DASHBOARD AND MAIN ROUTES
# ============================================================================

@app.route('/')
@login_required
def index():
    """
    Main dashboard with system overview and statistics
    
    Displays comprehensive system statistics including cylinder inventory,
    customer counts, utilization rates, and operational metrics. Provides
    quick access to key system information for all user roles.
    
    Features:
    - Total customers and cylinders count
    - Cylinder status breakdown (available, rented, maintenance)
    - Utilization rate calculation
    - System efficiency metrics
    - Growth rate and operational days
    - Role-based information display
    
    Returns:
        Dashboard template with comprehensive system statistics
    """
    # Get actual total counts from PostgreSQL database directly
    with CustomerService() as customer_service:
        customers, _ = customer_service.get_all(page=1, per_page=10000)
        customers_data, total_customers = customer_service.get_all(page=1, per_page=1)
    
    with CylinderService() as cylinder_service:
        # Get counts directly from database using SQL for accuracy
        from sqlalchemy import text
        
        # Get total cylinder count
        result = cylinder_service.db.execute(text("SELECT COUNT(*) FROM cylinders")).scalar()
        total_cylinders = result or 0
        
        # Get count by status (case-insensitive)
        available_result = cylinder_service.db.execute(text("SELECT COUNT(*) FROM cylinders WHERE LOWER(status) = 'available'")).scalar()
        available_cylinders = available_result or 0
        
        rented_result = cylinder_service.db.execute(text("SELECT COUNT(*) FROM cylinders WHERE LOWER(status) = 'rented'")).scalar()
        rented_cylinders = rented_result or 0
        
        maintenance_result = cylinder_service.db.execute(text("SELECT COUNT(*) FROM cylinders WHERE LOWER(status) = 'maintenance'")).scalar()
        maintenance_cylinders = maintenance_result or 0
        
        # Get customer rental counts for top customer calculation
        customer_rentals = {}
        rental_results = cylinder_service.db.execute(text("SELECT rented_to, COUNT(*) as count FROM cylinders WHERE status = 'rented' AND rented_to IS NOT NULL GROUP BY rented_to")).fetchall()
        for row in rental_results:
            customer_rentals[row[0]] = row[1]
    
    utilization_rate = round((rented_cylinders / total_cylinders * 100) if total_cylinders > 0 else 0)
    
    top_customer_count = max(customer_rentals.values()) if customer_rentals else 0
    
    # Calculate average rental days (mock data)
    import random
    avg_rental_days = random.randint(7, 30)
    
    # Calculate efficiency score (based on utilization and availability)
    efficiency_score = min(10, round((utilization_rate + (available_cylinders / total_cylinders * 100 if total_cylinders > 0 else 0)) / 20))
    
    # Days since first customer/cylinder - get from PostgreSQL
    from datetime import datetime
    
    days_active = 1
    try:
        with CustomerService() as customer_service:
            customers, _ = customer_service.get_all(page=1, per_page=10000)
        with CustomerService() as customer_service:
            customers, _ = customer_service.get_all(page=1, per_page=10000)
            oldest_customer = customer_service.db.execute(
                text("SELECT MIN(created_at) FROM customers WHERE created_at IS NOT NULL")
            ).scalar()
            
            if oldest_customer:
                days_active = (datetime.now().date() - oldest_customer.date()).days + 1
    except:
        # Fallback to 1 day if no data available
        days_active = 1
    
    # Growth rate (mock calculation)
    growth_rate = random.randint(5, 25)
    
    stats = {
        'total_customers': total_customers,
        'total_cylinders': total_cylinders,
        'available_cylinders': available_cylinders,
        'rented_cylinders': rented_cylinders,
        'maintenance_cylinders': maintenance_cylinders,
        'utilization_rate': utilization_rate,
        'top_customer_count': top_customer_count,
        'avg_rental_days': avg_rental_days,
        'efficiency_score': efficiency_score,
        'days_active': days_active,
        'growth_rate': growth_rate
    }
    
    return render_template('index.html', stats=stats)

@app.route('/metrics')
@login_required
def metrics():
    """Metrics and analytics page"""
    with CustomerService() as customer_service:
        customers, _ = customer_service.get_all(page=1, per_page=10000)
    with CylinderService() as cylinder_service:
        cylinders, _ = cylinder_service.get_all(page=1, per_page=1000)
    
    # Get cylinder status counts
    available_cylinders = len([c for c in cylinders if c.get('status', '').lower() == 'available'])
    rented_cylinders = len([c for c in cylinders if c.get('status', '').lower() == 'rented'])
    maintenance_cylinders = len([c for c in cylinders if c.get('status', '').lower() == 'maintenance'])
    
    # Calculate fun metrics
    total_cylinders = len(cylinders)
    utilization_rate = round((rented_cylinders / total_cylinders * 100) if total_cylinders > 0 else 0)
    
    # Find top customer (most rentals)
    customer_rentals = {}
    for cylinder in cylinders:
        if cylinder.get('rented_to'):
            customer_id = cylinder['rented_to']
            customer_rentals[customer_id] = customer_rentals.get(customer_id, 0) + 1
    
    top_customer_count = max(customer_rentals.values()) if customer_rentals else 0
    
    # Calculate average rental days (mock data)
    import random
    avg_rental_days = random.randint(7, 30)
    
    # Calculate efficiency score (based on utilization and availability)
    efficiency_score = min(10, round((utilization_rate + (available_cylinders / total_cylinders * 100 if total_cylinders > 0 else 0)) / 20))
    
    # Days since first customer/cylinder - get from PostgreSQL
    from datetime import datetime
    
    days_active = 1
    try:
        with CustomerService() as customer_service:
            customers, _ = customer_service.get_all(page=1, per_page=10000)
        with CustomerService() as customer_service:
            customers, _ = customer_service.get_all(page=1, per_page=10000)
            oldest_customer = customer_service.db.execute(
                text("SELECT MIN(created_at) FROM customers WHERE created_at IS NOT NULL")
            ).scalar()
            
            if oldest_customer:
                days_active = (datetime.now().date() - oldest_customer.date()).days + 1
    except:
        # Fallback to 1 day if no data available
        days_active = 1
    
    # Growth rate (mock calculation)
    growth_rate = random.randint(5, 25)
    
    stats = {
        'total_customers': len(customers),
        'total_cylinders': total_cylinders,
        'available_cylinders': available_cylinders,
        'rented_cylinders': rented_cylinders,
        'maintenance_cylinders': maintenance_cylinders,
        'utilization_rate': utilization_rate,
        'top_customer_count': top_customer_count,
        'avg_rental_days': avg_rental_days,
        'efficiency_score': efficiency_score,
        'days_active': days_active,
        'growth_rate': growth_rate
    }
    
    return render_template('metrics.html', stats=stats)

@app.route('/send_admin_stats', methods=['POST'])
@login_required
@admin_required
def send_admin_stats():
    """Send admin statistics via email"""
    if not EMAIL_AVAILABLE or not email_service:
        flash('Email service not available', 'error')
        return redirect(url_for('metrics'))
    
    email = request.form.get('email', '').strip()
    if not email:
        flash('Please enter a valid email address', 'error')
        return redirect(url_for('metrics'))
    
    # Get current stats
    with CustomerService() as customer_service:
        customers, _ = customer_service.get_all(page=1, per_page=10000)
    with CylinderService() as cylinder_service:
        cylinders, _ = cylinder_service.get_all(page=1, per_page=1000)
    
    # Get cylinder status counts
    available_cylinders = len([c for c in cylinders if c.get('status', '').lower() == 'available'])
    rented_cylinders = len([c for c in cylinders if c.get('status', '').lower() == 'rented'])
    maintenance_cylinders = len([c for c in cylinders if c.get('status', '').lower() == 'maintenance'])
    
    # Calculate metrics
    total_cylinders = len(cylinders)
    utilization_rate = round((rented_cylinders / total_cylinders * 100) if total_cylinders > 0 else 0)
    
    # Find top customer (most rentals)
    customer_rentals = {}
    for cylinder in cylinders:
        if cylinder.get('rented_to'):
            customer_id = cylinder['rented_to']
            customer_rentals[customer_id] = customer_rentals.get(customer_id, 0) + 1
    
    top_customer_count = max(customer_rentals.values()) if customer_rentals else 0
    
    # Calculate efficiency score (based on utilization and availability)
    efficiency_score = min(10, round((utilization_rate + (available_cylinders / total_cylinders * 100 if total_cylinders > 0 else 0)) / 20))
    
    # Days since first customer/cylinder - get from PostgreSQL
    from datetime import datetime
    
    days_active = 1
    try:
        with CustomerService() as customer_service:
            customers, _ = customer_service.get_all(page=1, per_page=10000)
        with CustomerService() as customer_service:
            customers, _ = customer_service.get_all(page=1, per_page=10000)
            oldest_customer = customer_service.db.execute(
                text("SELECT MIN(created_at) FROM customers WHERE created_at IS NOT NULL")
            ).scalar()
            
            if oldest_customer:
                days_active = (datetime.now().date() - oldest_customer.date()).days + 1
    except:
        # Fallback to 1 day if no data available
        days_active = 1
    
    stats = {
        'total_customers': len(customers),
        'total_cylinders': total_cylinders,
        'available_cylinders': available_cylinders,
        'rented_cylinders': rented_cylinders,
        'maintenance_cylinders': maintenance_cylinders,
        'utilization_rate': utilization_rate,
        'top_customer_count': top_customer_count,
        'efficiency_score': efficiency_score,
        'days_active': days_active
    }
    
    # Send email
    success = email_service.send_admin_stats(email, stats)
    
    if success:
        flash(f'Statistics sent successfully to {email}', 'success')
    else:
        flash('Failed to send email. Please check your email configuration.', 'error')
    
    return redirect(url_for('metrics'))

@app.route('/test_email', methods=['POST'])
@login_required
@admin_required
def test_email():
    """Send a test email to verify configuration"""
    if not EMAIL_AVAILABLE or not email_service:
        flash('Email service not available', 'error')
        return redirect(url_for('metrics'))
    
    email = request.form.get('test_email', '').strip()
    if not email:
        flash('Please enter a valid email address', 'error')
        return redirect(url_for('metrics'))
    
    success = email_service.send_test_email(email)
    
    if success:
        flash(f'Test email sent successfully to {email}', 'success')
    else:
        flash('Failed to send test email. Please check your email configuration.', 'error')
    
    return redirect(url_for('metrics'))

# Customer routes
@app.route('/customers')
@login_required
def customers():
    """Display all customers with search functionality and pagination"""
    search_query = request.args.get('search', '')
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 25))
    
    # Use a single optimized SQL query to get customers with their active dispatch counts
    from sqlalchemy import text
    with CustomerService() as customer_service:
        customers, _ = customer_service.get_all(page=1, per_page=10000)
    with CustomerService() as customer_service:
        customers, _ = customer_service.get_all(page=1, per_page=10000)
        if search_query:
            sql_query = """
            SELECT c.id, c.customer_no, c.customer_name, c.customer_email, c.customer_phone, 
                   c.customer_address, c.customer_city, c.customer_state, c.created_at,
                   COUNT(cyl.id) as active_dispatches
            FROM customers c 
            LEFT JOIN cylinders cyl ON cyl.rented_to = c.id AND cyl.status IN ('rented', 'dispatched')
            WHERE LOWER(c.customer_name) LIKE LOWER(:search) 
               OR LOWER(c.customer_no) LIKE LOWER(:search)
               OR LOWER(c.customer_phone) LIKE LOWER(:search)
            GROUP BY c.id, c.customer_no, c.customer_name, c.customer_email, c.customer_phone, 
                     c.customer_address, c.customer_city, c.customer_state, c.created_at
            ORDER BY active_dispatches DESC
            LIMIT :limit OFFSET :offset
            """
            search_param = f'%{search_query}%'
            result = customer_service.db.execute(
                text(sql_query), 
                {'search': search_param, 'limit': per_page, 'offset': (page - 1) * per_page}
            )
            
            # Count total for pagination
            count_query = """
            SELECT COUNT(DISTINCT c.id)
            FROM customers c
            WHERE LOWER(c.customer_name) LIKE LOWER(:search) 
               OR LOWER(c.customer_no) LIKE LOWER(:search)
               OR LOWER(c.customer_phone) LIKE LOWER(:search)
            """
            total_result = customer_service.db.execute(text(count_query), {'search': search_param})
            total_customers = total_result.scalar()
        else:
            sql_query = """
            SELECT c.id, c.customer_no, c.customer_name, c.customer_email, c.customer_phone, 
                   c.customer_address, c.customer_city, c.customer_state, c.created_at,
                   COUNT(cyl.id) as active_dispatches
            FROM customers c 
            LEFT JOIN cylinders cyl ON cyl.rented_to = c.id AND cyl.status IN ('rented', 'dispatched')
            GROUP BY c.id, c.customer_no, c.customer_name, c.customer_email, c.customer_phone, 
                     c.customer_address, c.customer_city, c.customer_state, c.created_at
            ORDER BY active_dispatches DESC
            LIMIT :limit OFFSET :offset
            """
            result = customer_service.db.execute(
                text(sql_query), 
                {'limit': per_page, 'offset': (page - 1) * per_page}
            )
            
            # Count total for pagination
            count_result = customer_service.db.execute(text("SELECT COUNT(*) FROM customers"))
            total_customers = count_result.scalar()
        
        # Convert results to dictionaries
        customers_paginated = []
        for row in result:
            # Clean phone number - convert empty/0.0 values to None
            phone = row.customer_phone
            if phone in ['0.0', '0', '', None]:
                phone = None
                
            # Handle created_at properly - it might be a string or datetime object
            created_at_str = None
            if row.created_at:
                if hasattr(row.created_at, 'isoformat'):
                    # It's a datetime object
                    created_at_str = row.created_at.isoformat()
                else:
                    # It's already a string
                    created_at_str = str(row.created_at)
                    
            customer_dict = {
                'id': row.id,
                'customer_no': row.customer_no,
                'customer_name': row.customer_name,
                'customer_email': row.customer_email,
                'customer_phone': phone,
                'customer_address': row.customer_address,
                'customer_city': row.customer_city,  
                'customer_state': row.customer_state,
                'created_at': created_at_str,
                'active_dispatches': row.active_dispatches,
                'rental_count': row.active_dispatches,
                'rented_cylinders': []  # We'll populate this only when needed
            }
            customers_paginated.append(customer_dict)
    
    # Calculate pagination info 
    total_pages = (total_customers + per_page - 1) // per_page
    has_prev = page > 1
    has_next = page < total_pages
    
    pagination_info = {
        'page': page,
        'per_page': per_page,
        'total': total_customers,
        'total_pages': total_pages,
        'has_prev': has_prev,
        'has_next': has_next,
        'prev_num': page - 1 if has_prev else None,
        'next_num': page + 1 if has_next else None,
        'start_index': ((page - 1) * per_page) + 1 if customers_paginated else 0,
        'end_index': min(page * per_page, total_customers)
    }
    
    return render_template('customers.html', 
                          customers=customers_paginated, 
                          search_query=search_query,
                          pagination=pagination_info)

@app.route('/customer/<customer_id>/details')
@login_required
def customer_details(customer_id):
    """Display detailed information for a specific customer with rental history tabs"""
    with CustomerService() as customer_service:
        customer_obj = customer_service.get_by_id(customer_id)
        
        if not customer_obj:
            flash('Customer not found', 'error')
            return redirect(url_for('customers'))
        
        # Convert customer to dict within the session to avoid detached instance errors
        created_at_str = None
        if customer_obj.created_at:
            if hasattr(customer_obj.created_at, 'isoformat'):
                created_at_str = customer_obj.created_at.isoformat()
            else:
                created_at_str = str(customer_obj.created_at)
                    
        customer = {
            'id': customer_obj.id,
            'customer_no': customer_obj.customer_no,
            'customer_name': customer_obj.customer_name,
            'customer_email': customer_obj.customer_email,
            'customer_phone': customer_obj.customer_phone,
            'customer_address': customer_obj.customer_address,
            'customer_city': customer_obj.customer_city,
            'customer_state': customer_obj.customer_state,
            'created_at': created_at_str
        }
    
    # Get rental history (active and past) from PostgreSQL
    from db_service import RentalHistoryService
    
    # Get active rentals directly from cylinder service (already returns dictionaries)
    with CylinderService() as cylinder_service:
        active_rentals_dict = cylinder_service.get_by_customer(customer['id'])
    
    # Get customer's rental history from rental history service
    with RentalHistoryService() as history_service:
        history_result = history_service.get_customer_history(customer['id'])
        past_transactions = history_result['past']
    
    # Past transactions are already dictionaries from the service
    past_transactions_dict = past_transactions[:50]  # Limit to recent 50 transactions for performance
    

    
    history_data = {
        'active': active_rentals_dict,
        'past': past_transactions_dict
    }
    
    # Get tab parameter
    tab = request.args.get('tab', 'active')
    
    # Get pagination parameters
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 25))
    
    # Select data based on tab
    if tab == 'past':
        cylinders_data = history_data['past']
    else:
        cylinders_data = history_data['active']
    
    # Pagination
    total_cylinders = len(cylinders_data)
    start = (page - 1) * per_page
    end = start + per_page
    cylinders_paginated = cylinders_data[start:end]
    
    # Calculate pagination info
    total_pages = (total_cylinders + per_page - 1) // per_page
    has_prev = page > 1
    has_next = page < total_pages
    
    pagination_info = {
        'page': page,
        'per_page': per_page,
        'total': total_cylinders,
        'total_pages': total_pages,
        'has_prev': has_prev,
        'has_next': has_next,
        'prev_num': page - 1 if has_prev else None,
        'next_num': page + 1 if has_next else None,
        'start_index': start + 1 if cylinders_paginated else 0,
        'end_index': min(end, total_cylinders)
    }
    
    # Calculate summary statistics
    active_count = len(history_data['active'])
    past_count = len(history_data['past'])
    
    if tab == 'active' and history_data['active']:
        avg_rental_days = sum(c.get('rental_days', 0) for c in history_data['active']) // active_count
        long_term_count = len([c for c in history_data['active'] if c.get('rental_days', 0) > 90])
    elif tab == 'past' and history_data['past']:
        avg_rental_days = sum(c.get('rental_days', 0) for c in history_data['past']) // past_count
        long_term_count = len([c for c in history_data['past'] if c.get('rental_days', 0) > 90])
    else:
        avg_rental_days = 0
        long_term_count = 0
    
    return render_template('customer_details.html', 
                         customer=customer, 
                         cylinders_data=cylinders_paginated,
                         active_count=active_count,
                         past_count=past_count,
                         current_tab=tab,
                         avg_rental_days=avg_rental_days,
                         long_term_count=long_term_count,
                         pagination=pagination_info)

@app.route('/customer/<customer_id>/monthly_history')
@login_required
def customer_monthly_history(customer_id):
    """Display customer's monthly activity history"""
    from datetime import datetime
    
    with CustomerService() as customer_service:
        customers, _ = customer_service.get_all(page=1, per_page=10000)
    with CustomerService() as customer_service:
        customers, _ = customer_service.get_all(page=1, per_page=10000)
    
    if not customer_obj:
        flash('Customer not found', 'error')
        return redirect(url_for('customers'))
    
    # Convert customer to dict
    if hasattr(customer_obj, 'id'):
        customer = {
            'id': customer_obj.id,
            'customer_no': customer_obj.customer_no,
            'customer_name': customer_obj.customer_name,
            'customer_email': customer_obj.customer_email,
            'customer_phone': customer_obj.customer_phone,
            'customer_address': customer_obj.customer_address,
            'customer_city': customer_obj.customer_city,
            'customer_state': customer_obj.customer_state
        }
    else:
        customer = customer_obj
    
    # Get year parameter
    year = request.args.get('year', type=int)
    if not year:
        year = datetime.now().year
    
    # Get monthly history data
    with RentalHistoryService() as history_service:
        history_data = history_service.get_customer_monthly_history(customer['id'], year)
    
    return render_template('customer_monthly_history.html',
                          customer=customer,
                          monthly_data=history_data['monthly_data'],
                          available_years=history_data['available_years'],
                          selected_year=year)

@app.route('/customers/add', methods=['GET', 'POST'])
@admin_or_user_can_edit
def add_customer():
    """Add new customer"""
    if request.method == 'POST':
        # Validate required fields for new customer structure
        # Required: customer_no, customer_name, customer_address, customer_city, customer_state, customer_phone
        # Optional: customer_apgst, customer_cst
        required_fields = ['customer_no', 'customer_name', 'customer_address', 'customer_city', 'customer_state', 'customer_phone']
        customer_data = {}
        
        for field in required_fields:
            value = request.form.get(field, '').strip()
            if not value:
                # Create user-friendly field names for error messages
                display_name = field.replace('customer_', '').replace('_', ' ').title()
                flash(f'{display_name} is required', 'error')
                return render_template('add_customer.html')
            customer_data[field] = value
        
        # Add optional fields
        customer_data['customer_apgst'] = request.form.get('customer_apgst', '').strip()
        customer_data['customer_cst'] = request.form.get('customer_cst', '').strip()
        
        # Add optional email field
        customer_data['customer_email'] = request.form.get('customer_email', '').strip()
        
        try:
            with CustomerService() as customer_service:
                new_customer = customer_service.create(customer_data)
            if new_customer:
                customer_name = customer_data.get('customer_name', 'Unknown')
                customer_id = new_customer.id if hasattr(new_customer, 'id') else 'N/A'
                flash(f'Customer {customer_name} added successfully with ID: {customer_id}', 'success')
                return redirect(url_for('customers'))
            else:
                flash('Error adding customer', 'error')
        except Exception as e:
            flash(f'Error adding customer: {str(e)}', 'error')
    
    return render_template('add_customer.html')

@app.route('/customers/edit/<customer_id>', methods=['GET', 'POST'])
@admin_or_user_can_edit
def edit_customer(customer_id):
    """Edit existing customer"""
    with CustomerService() as customer_service:
        customer = customer_service.get_by_id(customer_id)
    
    if not customer:
        flash('Customer not found', 'error')
        return redirect(url_for('customers'))
    
    # Convert customer to dictionary if it's a SQLAlchemy object
    if hasattr(customer, 'id'):
        customer_dict = {
            'id': customer.id,
            'customer_no': customer.customer_no,
            'customer_name': customer.customer_name,
            'customer_email': customer.customer_email or '',
            'customer_phone': customer.customer_phone,
            'customer_address': customer.customer_address,
            'customer_city': customer.customer_city,
            'customer_state': customer.customer_state,
            'customer_apgst': getattr(customer, 'customer_apgst', '') or '',
            'customer_cst': getattr(customer, 'customer_cst', '') or ''
        }
    else:
        customer_dict = customer
    
    if request.method == 'POST':
        # Validate required fields for new customer structure
        # Required: customer_no, customer_name, customer_address, customer_city, customer_state, customer_phone
        # Optional: customer_apgst, customer_cst, customer_email
        required_fields = ['customer_no', 'customer_name', 'customer_address', 'customer_city', 'customer_state', 'customer_phone']
        customer_data = {}
        
        for field in required_fields:
            value = request.form.get(field, '').strip()
            if not value:
                # Create user-friendly field names for error messages
                display_name = field.replace('customer_', '').replace('_', ' ').title()
                flash(f'{display_name} is required', 'error')
                return render_template('edit_customer.html', customer=customer_dict)
            customer_data[field] = value
        
        # Add optional fields
        customer_data['customer_email'] = request.form.get('customer_email', '').strip()
        customer_data['customer_apgst'] = request.form.get('customer_apgst', '').strip()
        customer_data['customer_cst'] = request.form.get('customer_cst', '').strip()
        
        try:
            with CustomerService() as customer_service:
                updated_customer = customer_service.update(customer_id, customer_data)
            if updated_customer:
                flash(f'Customer {customer_data["customer_name"]} updated successfully', 'success')
                return redirect(url_for('customers'))
            else:
                flash('Error updating customer', 'error')
        except Exception as e:
            flash(f'Error updating customer: {str(e)}', 'error')
    
    return render_template('edit_customer.html', customer=customer_dict)

@app.route('/customers/delete/<customer_id>', methods=['POST'])
@admin_or_user_can_edit
def delete_customer(customer_id):
    """Delete customer"""
    try:
        with CustomerService() as customer_service:
            deleted = customer_service.delete(customer_id)
        if deleted:
            flash('Customer deleted successfully', 'success')
        else:
            flash('Customer not found', 'error')
    except Exception as e:
        flash(f'Error deleting customer: {str(e)}', 'error')
    
    return redirect(url_for('customers'))

# Rental History routes
@app.route('/rental_history')
@login_required
def rental_history():
    """Display rental history"""
    from db_service import RentalHistoryService
    
    # Only cleanup if explicitly requested
    cleanup_requested = request.args.get('cleanup', False)
    removed_count = 0
    
    if cleanup_requested and session.get('role') == 'admin':
        with RentalHistoryService() as service:
            removed_count = service.cleanup_old_records()
        
        if removed_count > 0:
            flash(f'Removed {removed_count} records older than 6 months', 'info')
    
    # Get pagination parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    
    # Limit per_page to reasonable values
    per_page = min(max(per_page, 10), 200)
    
    search_query = request.args.get('search', '')
    customer_filter = request.args.get('customer', '')
    
    # Use RentalHistory service to get return records and convert within session  
    with RentalHistoryService() as service:
        all_transactions, total_count = service.get_all(page=1, per_page=5000)  # Get larger batch
        print(f"Debug: Retrieved {len(all_transactions)} transactions out of {total_count} total")
        
        # Convert SQLAlchemy objects to dicts immediately within the active session
        transaction_dicts = []
        for t in all_transactions:
            if isinstance(t, dict):  # Already a dict
                transaction_dicts.append(t)
            else:  # SQLAlchemy object - convert within active session
                try:
                    # Access all attributes while session is active
                    transaction_dict = {
                        'id': t.id if hasattr(t, 'id') else '',
                        'customer_name': t.customer_name if hasattr(t, 'customer_name') and t.customer_name else '',
                        'cylinder_custom_id': t.cylinder_custom_id if hasattr(t, 'cylinder_custom_id') and t.cylinder_custom_id else '',
                        'customer_no': t.customer_no if hasattr(t, 'customer_no') and t.customer_no else '',
                        'return_date': t.return_date.isoformat() if hasattr(t, 'return_date') and t.return_date else '',
                        'dispatch_date': t.dispatch_date.isoformat() if hasattr(t, 'dispatch_date') and t.dispatch_date else '',
                        'rental_days': t.rental_days if hasattr(t, 'rental_days') and t.rental_days else 0,
                        'cylinder_type': t.cylinder_type if hasattr(t, 'cylinder_type') and t.cylinder_type else '',
                        'cylinder_size': t.cylinder_size if hasattr(t, 'cylinder_size') and t.cylinder_size else '',
                        'customer_phone': t.customer_phone if hasattr(t, 'customer_phone') and t.customer_phone else '',
                        'customer_address': t.customer_address if hasattr(t, 'customer_address') and t.customer_address else '',
                        'location': t.location if hasattr(t, 'location') and t.location else '',
                        'status': t.status if hasattr(t, 'status') and t.status else ''
                    }
                    transaction_dicts.append(transaction_dict)
                except Exception as e:
                    app.logger.error(f"Error converting transaction: {str(e)}")
                    # Skip problematic transactions but continue processing
                    continue
    
    # Use the converted transactions
    all_transactions = transaction_dicts
    
    # Apply search filter
    if search_query:
        all_transactions = [t for t in all_transactions 
                          if search_query.lower() in t.get('customer_name', '').lower() or
                             search_query.lower() in t.get('cylinder_custom_id', '').lower() or
                             search_query.lower() in t.get('customer_no', '').lower()]
    
    # Apply customer filter  
    if customer_filter:
        all_transactions = [t for t in all_transactions 
                          if t.get('customer_no', '').upper() == customer_filter.upper()]
    
    # Sort by return date (most recent first)
    all_transactions.sort(key=lambda x: x.get('return_date', '') or x.get('date_returned', ''), reverse=True)
    
    # Calculate pagination
    total_transactions = len(all_transactions)
    start = (page - 1) * per_page
    end = start + per_page
    transactions_paginated = all_transactions[start:end]
    
    # Calculate pagination info
    total_pages = (total_transactions + per_page - 1) // per_page
    has_prev = page > 1
    has_next = page < total_pages
    
    pagination_info = {
        'page': page,
        'per_page': per_page,
        'total': total_transactions,
        'total_pages': total_pages,
        'has_prev': has_prev,
        'has_next': has_next,
        'prev_num': page - 1 if has_prev else None,
        'next_num': page + 1 if has_next else None,
        'start_index': start + 1 if transactions_paginated else 0,
        'end_index': min(end, total_transactions)
    }
    
    # Get unique customers for filter dropdown
    unique_customers = list(set((t.get('customer_no', ''), t.get('customer_name', '')) 
                               for t in all_transactions if t.get('customer_no')))
    unique_customers.sort(key=lambda x: x[1])  # Sort by customer name
    
    return render_template('rental_history.html',
                         transactions=transactions_paginated,
                         pagination=pagination_info,
                         search_query=search_query,
                         customer_filter=customer_filter,
                         unique_customers=unique_customers,
                         total_transactions=total_transactions)

# Cylinder routes
@app.route('/cylinders')
@login_required
def cylinders():
    """List all cylinders with search, filter functionality, and pagination"""

    # Get pagination parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)  # Default 50 cylinders per page
    
    # Limit per_page to reasonable values
    per_page = min(max(per_page, 10), 200)  # Between 10 and 200 items per page
    
    search_query = request.args.get('search', '')
    status_filter = request.args.get('status', '')
    customer_filter = request.args.get('customer', '')
    type_filter = request.args.get('type_filter', '')
    rental_duration_filter = request.args.get('rental_duration', '')
    
    # Get all cylinders with search and pagination using PostgreSQL service
    with CylinderService() as cylinder_service:
        cylinders_list, total_cylinders = cylinder_service.get_all(
            search_query=search_query,
            page=page,
            per_page=per_page,
            filter_type=type_filter,
            filter_status=status_filter,
            rental_duration_filter=rental_duration_filter,
            customer_filter=customer_filter
        )
    
    # PostgreSQL model already returns dictionaries with calculated fields
    paginated_cylinders = cylinders_list
    
    # Calculate pagination info
    total_pages = (total_cylinders + per_page - 1) // per_page
    has_prev = page > 1
    has_next = page < total_pages
    prev_page = page - 1 if has_prev else None
    next_page = page + 1 if has_next else None
    
    # Create pagination object for template
    pagination = {
        'page': page,
        'per_page': per_page,
        'total': total_cylinders,
        'total_pages': total_pages,
        'has_prev': has_prev,
        'has_next': has_next,
        'prev_page': prev_page,
        'next_page': next_page,
        'pages': list(range(max(1, page - 2), min(total_pages + 1, page + 3)))  # Show 5 pages around current
    }
    
    # Get all customers for the filter dropdown using PostgreSQL service
    # Convert to dictionaries while session is still active
    customers = []
    with CustomerService() as customer_service:
        customers_list, _ = customer_service.get_all(page=1, per_page=10000)
        
        # Convert customers to dictionaries while session is active
        for customer in customers_list:
            try:
                # Try to access as SQLAlchemy object first
                customers.append({
                    'id': customer.id,
                    'customer_name': customer.customer_name,
                    'customer_no': customer.customer_no,
                    'customer_phone': customer.customer_phone,
                    'customer_email': customer.customer_email or '',
                    'customer_city': customer.customer_city,
                    'customer_state': customer.customer_state
                })
            except (AttributeError, Exception):
                # If it's already a dictionary or any other issue, use as-is
                customers.append(customer)

    return render_template('cylinders.html', 
                         cylinders=paginated_cylinders, 
                         customers=customers,
                         search_query=search_query,
                         status_filter=status_filter,
                         customer_filter=customer_filter,
                         type_filter=type_filter,
                         pagination=pagination,
                         rental_duration_filter=rental_duration_filter,
                         )

@app.route('/cylinders/<cylinder_id>/details')
@login_required
def cylinder_details(cylinder_id):
    """Display detailed information for a specific cylinder"""
    # Convert SQLAlchemy object to dictionary while session is active
    cylinder = None
    with CylinderService() as cylinder_service:
        cylinder_obj = cylinder_service.get_by_id(cylinder_id)
        if not cylinder_obj:
            flash('Cylinder not found', 'error')
            return redirect(url_for('cylinders'))
        
        # Convert SQLAlchemy object to dictionary while session is active
        cylinder = {
            'id': cylinder_obj.id,
            'custom_id': cylinder_obj.custom_id or '',
            'serial_number': cylinder_obj.serial_number or '',
            'type': cylinder_obj.type or 'Medical Oxygen',
            'size': cylinder_obj.size or '40L',
            'status': cylinder_obj.status or 'available',
            'location': cylinder_obj.location or 'Warehouse',
            'pressure': getattr(cylinder_obj, 'pressure', ''),
            'last_inspection': getattr(cylinder_obj, 'last_inspection', None),
            'next_inspection': getattr(cylinder_obj, 'next_inspection', None),
            'notes': getattr(cylinder_obj, 'notes', ''),
            'rented_to': cylinder_obj.rented_to,
            'customer_name': cylinder_obj.customer_name or '',
            'customer_no': cylinder_obj.customer_no or '',
            'customer_email': getattr(cylinder_obj, 'customer_email', ''),
            'customer_phone': getattr(cylinder_obj, 'customer_phone', ''),
            'customer_city': getattr(cylinder_obj, 'customer_city', ''),
            'customer_state': getattr(cylinder_obj, 'customer_state', ''),
            'date_borrowed': cylinder_obj.date_borrowed,
            'rental_date': cylinder_obj.rental_date,
            'date_returned': cylinder_obj.date_returned,
            'created_at': cylinder_obj.created_at,
            'updated_at': cylinder_obj.updated_at
        }
    
    # Add display ID (custom_id if available, otherwise generated serial)
    cylinder['display_id'] = cylinder['custom_id'] or cylinder['serial_number'] or f"ID-{cylinder['id'][:8]}"
    
    # Add rental days calculation
    if cylinder['rental_date']:
        try:
            rental_days = (datetime.utcnow() - cylinder['rental_date']).days
            cylinder['rental_days'] = rental_days
            cylinder['rental_months'] = rental_days // 30
        except:
            cylinder['rental_days'] = 0
            cylinder['rental_months'] = 0
    else:
        cylinder['rental_days'] = 0
        cylinder['rental_months'] = 0
    
    # Get rental history for this cylinder
    rental_history = []
    try:
        with RentalHistoryService() as history_service:
            rental_history = history_service.get_by_cylinder(cylinder['id'])
    except Exception as e:
        print(f"Error getting rental history: {e}")
    
    return render_template('cylinder_details.html', cylinder=cylinder, rental_history=rental_history)

@app.route('/cylinders/add', methods=['GET', 'POST'])
@admin_or_user_can_edit
def add_cylinder():
    """Add new cylinder"""
    if request.method == 'POST':
        # Validate required fields - custom_id is now REQUIRED
        required_fields = ['custom_id', 'type', 'size', 'status', 'location']
        cylinder_data = {}
        
        for field in required_fields:
            value = request.form.get(field, '').strip()
            if not value:
                field_display = 'ID' if field == 'custom_id' else field.replace('_', ' ').title()
                flash(f'{field_display} is required', 'error')
                with CustomerService() as customer_service:
                    customers, _ = customer_service.get_all(page=1, per_page=10000)
                with CustomerService() as customer_service:
                    customers, _ = customer_service.get_all(page=1, per_page=10000)
                return render_template('add_cylinder.html', customers=customers, today_date=datetime.now().strftime('%Y-%m-%d'))
            cylinder_data[field] = value
        
        # Add optional fields - serial number is now optional
        # (No need to set custom_id here since it's now in required_fields)
        cylinder_data['pressure'] = request.form.get('pressure', '').strip()
        cylinder_data['last_inspection'] = request.form.get('last_inspection', '').strip()
        cylinder_data['next_inspection'] = request.form.get('next_inspection', '').strip()
        cylinder_data['notes'] = request.form.get('notes', '').strip()
        
        # Validate custom_id uniqueness (now required)
        with CylinderService() as cylinder_service:
            existing_cylinders, _ = cylinder_service.get_all(page=1, per_page=1000)
            for existing in existing_cylinders:
                existing_custom_id = existing.custom_id if hasattr(existing, 'custom_id') else existing.get('custom_id', '')
                if existing_custom_id == cylinder_data['custom_id']:
                    flash(f'ID "{cylinder_data["custom_id"]}" is already in use. Please choose a different one.', 'error')
                    with CustomerService() as customer_service:
                        customers, _ = customer_service.get_all(page=1, per_page=10000)
                    with CustomerService() as customer_service:
                        customers, _ = customer_service.get_all(page=1, per_page=10000)
                    return render_template('add_cylinder.html', customers=customers, today_date=datetime.now().strftime('%Y-%m-%d'))
        
        # Handle customer assignment for rented cylinders
        rented_to = request.form.get('rented_to', '').strip()
        if cylinder_data['status'].lower() == 'rented':
            if not rented_to:
                flash('Customer selection is required when status is "Rented"', 'error')
                with CustomerService() as customer_service:
                    customers, _ = customer_service.get_all(page=1, per_page=10000)
                with CustomerService() as customer_service:
                    customers, _ = customer_service.get_all(page=1, per_page=10000)
                return render_template('add_cylinder.html', customers=customers)
            
            # Verify customer exists
            with CustomerService() as customer_service:
                customers, _ = customer_service.get_all(page=1, per_page=10000)
            with CustomerService() as customer_service:
                customers, _ = customer_service.get_all(page=1, per_page=10000)
                if not customer:
                    flash('Selected customer not found', 'error')
                    customers, _ = customer_service.get_all()
                    return render_template('add_cylinder.html', customers=customers)
            
            cylinder_data['rented_to'] = rented_to
            cylinder_data['customer_name'] = customer.get('customer_name', '')
            cylinder_data['customer_email'] = customer.get('customer_email', '')
            
            # Handle rental date from form or use current date
            rental_date = request.form.get('rental_date', '').strip()
            from datetime import datetime
            if rental_date:
                # Convert date string to ISO format
                try:
                    date_obj = datetime.strptime(rental_date, '%Y-%m-%d')
                    cylinder_data['date_borrowed'] = date_obj.isoformat()
                    # rental_date field removed - using date_borrowed instead
                except ValueError:
                    # Fallback to current date if invalid format
                    cylinder_data['date_borrowed'] = datetime.now().isoformat()
                    # rental_date field removed - using date_borrowed instead
            else:
                cylinder_data['date_borrowed'] = datetime.now().isoformat()
        
        try:
            with CylinderService() as cylinder_service:
                new_cylinder = cylinder_service.create(cylinder_data)
            flash(f'Cylinder added successfully with ID: {new_cylinder.id}', 'success')
            return redirect(url_for('cylinders'))
        except Exception as e:
            flash(f'Error adding cylinder: {str(e)}', 'error')
    
    # Get all customers for the dropdown and today's date
    # Convert customers to dictionaries while session is active
    customers = []
    with CustomerService() as customer_service:
        customers_list, _ = customer_service.get_all()
        
        # Convert customers to dictionaries while session is active
        for customer in customers_list:
            try:
                customers.append({
                    'id': customer.id,
                    'customer_name': customer.customer_name,
                    'customer_no': customer.customer_no,
                    'customer_phone': customer.customer_phone,
                    'customer_email': customer.customer_email or '',
                    'customer_city': customer.customer_city,
                    'customer_state': customer.customer_state
                })
            except (AttributeError, Exception):
                customers.append(customer)
    
    from datetime import datetime
    today_date = datetime.now().strftime('%Y-%m-%d')
    return render_template('add_cylinder.html', customers=customers, today_date=today_date)

@app.route('/cylinders/edit/<cylinder_id>', methods=['GET', 'POST'])
@admin_or_user_can_edit
def edit_cylinder(cylinder_id):
    """Edit existing cylinder"""
    # Convert SQLAlchemy object to dictionary while session is active
    cylinder = None
    with CylinderService() as cylinder_service:
        cylinder_obj = cylinder_service.get_by_id(cylinder_id)
        if not cylinder_obj:
            flash('Cylinder not found', 'error')
            return redirect(url_for('cylinders'))
        
        # Convert SQLAlchemy object to dictionary while session is active
        cylinder = {
            'id': cylinder_obj.id,
            'custom_id': cylinder_obj.custom_id or '',
            'serial_number': cylinder_obj.serial_number or '',
            'type': cylinder_obj.type or 'Medical Oxygen',
            'size': cylinder_obj.size or '40L',
            'status': cylinder_obj.status or 'available',
            'location': cylinder_obj.location or 'Warehouse',
            'pressure': getattr(cylinder_obj, 'pressure', ''),
            'last_inspection': getattr(cylinder_obj, 'last_inspection', None),
            'next_inspection': getattr(cylinder_obj, 'next_inspection', None),
            'notes': getattr(cylinder_obj, 'notes', ''),
            'rented_to': cylinder_obj.rented_to,
            'customer_name': cylinder_obj.customer_name or '',
            'customer_no': cylinder_obj.customer_no or '',
            'date_borrowed': cylinder_obj.date_borrowed.isoformat() if cylinder_obj.date_borrowed else '',
            'date_returned': cylinder_obj.date_returned.isoformat() if cylinder_obj.date_returned else ''
        }
    
    if request.method == 'POST':
        # Validate required fields
        required_fields = ['type', 'size', 'status', 'location']
        cylinder_data = {}
        
        for field in required_fields:
            value = request.form.get(field, '').strip()
            if not value:
                flash(f'{field.replace("_", " ").title()} is required', 'error')
                # Get customers for dropdown while converting to dictionaries
                customers = []
                with CustomerService() as customer_service:
                    customers_list, _ = customer_service.get_all(page=1, per_page=10000)
                    for customer in customers_list:
                        try:
                            customers.append({
                                'id': customer.id,
                                'customer_name': customer.customer_name,
                                'customer_no': customer.customer_no
                            })
                        except:
                            customers.append(customer)
                return render_template('edit_cylinder.html', cylinder=cylinder, customers=customers)
            cylinder_data[field] = value
        
        # Add optional fields
        cylinder_data['custom_id'] = request.form.get('custom_id', '').strip()
        cylinder_data['pressure'] = request.form.get('pressure', '').strip()
        cylinder_data['last_inspection'] = request.form.get('last_inspection', '').strip()
        cylinder_data['next_inspection'] = request.form.get('next_inspection', '').strip()
        cylinder_data['notes'] = request.form.get('notes', '').strip()
        
        # Validate custom_id uniqueness if provided and different from current
        if cylinder_data['custom_id']:
            with CylinderService() as cylinder_service:
                existing_cylinders, _ = cylinder_service.get_all(page=1, per_page=1000)
            for existing in existing_cylinders:
                existing_custom_id = existing.custom_id if hasattr(existing, 'custom_id') else existing.get('custom_id', '')
                existing_id = existing.id if hasattr(existing, 'id') else existing.get('id', '')
                if existing_custom_id == cylinder_data['custom_id'] and existing_id != cylinder_id:
                    flash(f'Custom ID "{cylinder_data["custom_id"]}" is already in use. Please choose a different one.', 'error')
                    # Get customers for dropdown while converting to dictionaries
                    customers = []
                    with CustomerService() as customer_service:
                        customers_list, _ = customer_service.get_all(page=1, per_page=10000)
                        for customer in customers_list:
                            try:
                                customers.append({
                                    'id': customer.id,
                                    'customer_name': customer.customer_name,
                                    'customer_no': customer.customer_no,
                                    'customer_email': customer.customer_email or ''
                                })
                            except:
                                customers.append(customer)
                    return render_template('edit_cylinder.html', cylinder=cylinder, customers=customers)
        
        # Handle customer assignment for rented cylinders
        rented_to = request.form.get('rented_to', '').strip()
        if cylinder_data['status'].lower() == 'rented':
            if not rented_to:
                flash('Customer selection is required when status is "Rented"', 'error')
                # Get customers for dropdown while converting to dictionaries
                customers = []
                with CustomerService() as customer_service:
                    customers_list, _ = customer_service.get_all(page=1, per_page=10000)
                    for customer in customers_list:
                        try:
                            customers.append({
                                'id': customer.id,
                                'customer_name': customer.customer_name,
                                'customer_no': customer.customer_no,
                                'customer_email': customer.customer_email or ''
                            })
                        except:
                            customers.append(customer)
                return render_template('edit_cylinder.html', cylinder=cylinder, customers=customers)
            
            # Verify customer exists
            with CustomerService() as customer_service:
                customer = customer_service.get_by_id(rented_to)
            if not customer:
                flash('Selected customer not found', 'error')
                # Get customers for dropdown while converting to dictionaries
                customers = []
                with CustomerService() as customer_service:
                    customers_list, _ = customer_service.get_all(page=1, per_page=10000)
                    for customer in customers_list:
                        try:
                            customers.append({
                                'id': customer.id,
                                'customer_name': customer.customer_name,
                                'customer_no': customer.customer_no,
                                'customer_email': customer.customer_email or ''
                            })
                        except:
                            customers.append(customer)
                return render_template('edit_cylinder.html', cylinder=cylinder, customers=customers)
            
            cylinder_data['rented_to'] = rented_to
        else:
            # Clear customer assignment if not rented
            cylinder_data['rented_to'] = ''

        # Handle date tracking fields
        date_borrowed = request.form.get('date_borrowed', '').strip()
        date_returned = request.form.get('date_returned', '').strip()
        
        # Get current status and new status
        current_status = cylinder.get('status', '').lower()
        new_status = cylinder_data['status'].lower()
        
        # Auto-set dates based on status changes
        from datetime import datetime
        
        # If status is changing to 'rented', set date_borrowed
        if new_status == 'rented' and current_status != 'rented':
            if not date_borrowed:
                cylinder_data['date_borrowed'] = datetime.now().isoformat()
            else:
                # Convert datetime-local format to ISO format
                try:
                    dt = datetime.fromisoformat(date_borrowed)
                    cylinder_data['date_borrowed'] = dt.isoformat()
                except:
                    cylinder_data['date_borrowed'] = datetime.now().isoformat()
            # Clear return date when renting
            cylinder_data['date_returned'] = ''
            
        # If status is changing to 'available' from 'rented', set date_returned
        elif new_status == 'available' and current_status == 'rented':
            if not date_returned:
                cylinder_data['date_returned'] = datetime.now().isoformat()
            else:
                # Convert datetime-local format to ISO format
                try:
                    dt = datetime.fromisoformat(date_returned)
                    cylinder_data['date_returned'] = dt.isoformat()
                except:
                    cylinder_data['date_returned'] = datetime.now().isoformat()
        
        # If manually setting dates, convert them to proper format
        else:
            if date_borrowed:
                try:
                    dt = datetime.fromisoformat(date_borrowed)
                    cylinder_data['date_borrowed'] = dt.isoformat()
                except:
                    pass
            if date_returned:
                try:
                    dt = datetime.fromisoformat(date_returned)
                    cylinder_data['date_returned'] = dt.isoformat()
                except:
                    pass
        
        try:
            with CylinderService() as cylinder_service:
                updated_cylinder = cylinder_service.update(cylinder_id, cylinder_data)
            if updated_cylinder:
                display_id = cylinder_data.get('custom_id') or cylinder['custom_id'] or f"ID-{cylinder_id[:8]}"
                flash(f'Cylinder {display_id} updated successfully', 'success')
                return redirect(url_for('cylinders'))
            else:
                flash('Error updating cylinder', 'error')
        except Exception as e:
            flash(f'Error updating cylinder: {str(e)}', 'error')
    
    # Get all customers for the dropdown and add display ID
    # Convert customers to dictionaries while session is active
    customers = []
    with CustomerService() as customer_service:
        customers_list, _ = customer_service.get_all(page=1, per_page=10000)
        for customer in customers_list:
            try:
                customers.append({
                    'id': customer.id,
                    'customer_name': customer.customer_name,
                    'customer_no': customer.customer_no,
                    'customer_email': customer.customer_email or ''
                })
            except:
                customers.append(customer)
    
    cylinder['display_serial'] = cylinder.get('custom_id') or cylinder.get('serial_number') or f"ID-{cylinder['id'][:8]}"
    return render_template('edit_cylinder.html', cylinder=cylinder, customers=customers)

@app.route('/cylinders/delete/<cylinder_id>', methods=['POST'])
@admin_or_user_can_edit
def delete_cylinder(cylinder_id):
    """Delete cylinder"""
    try:
        with CylinderService() as cylinder_service:
            cylinder = cylinder_service.get_by_id(cylinder_id)
            if cylinder and cylinder_service.delete(cylinder_id):
                flash('Cylinder deleted successfully', 'success')
            else:
                flash('Cylinder not found', 'error')
    except Exception as e:
        flash(f'Error deleting cylinder: {str(e)}', 'error')
    
    return redirect(url_for('cylinders'))

# JSON Import routes
@app.route('/import/json', methods=['GET', 'POST'])
@login_required
def import_json():
    """JSON file import interface"""
    if request.method == 'POST':
        if 'json_file' not in request.files:
            flash('No file selected', 'error')
            return redirect(request.url)
        
        file = request.files['json_file']
        if file.filename == '':
            flash('No file selected', 'error')
            return redirect(request.url)
        
        if file and file.filename.lower().endswith('.json'):
            try:
                # Save uploaded file temporarily
                import tempfile
                import json
                
                # Read file content as bytes first
                file_content = file.read()
                
                # Try to decode and validate JSON
                try:
                    json_str = file_content.decode('utf-8')
                    json.loads(json_str)  # Validate JSON syntax
                except (UnicodeDecodeError, json.JSONDecodeError) as e:
                    flash(f'Invalid JSON file: {str(e)}', 'error')
                    return redirect(request.url)
                
                # Save to temporary file
                with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as temp_file:
                    temp_file.write(json_str)
                    temp_file_path = temp_file.name
                
                # Analyze the JSON file
                importer = JSONImporter()
                analysis = importer.analyze_json_file(temp_file_path)
                
                # Clean up temp file
                os.unlink(temp_file_path)
                
                if analysis['error']:
                    flash(f'Error analyzing file: {analysis["error"]}', 'error')
                    return redirect(request.url)
                
                # Store analysis in session for next step - compress large data
                # Don't store the full data in session to avoid size limits
                session_data = {
                    'data_type': analysis['data_type'],
                    'records': analysis['records'],
                    'fields': analysis['fields'],
                    'sample_record': analysis['sample_record'],
                    'error': analysis['error']
                }
                session['json_analysis'] = session_data
                session['json_filename'] = file.filename
                
                # Store the data separately in a temporary file for the next step
                import tempfile
                with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as data_file:
                    json.dump(analysis['data'], data_file)
                    session['json_data_path'] = data_file.name
                
                return redirect(url_for('map_json_fields'))
                
            except Exception as e:
                flash(f'Error processing file: {str(e)}', 'error')
                return redirect(request.url)
        else:
            flash('Please upload a valid JSON file', 'error')
            return redirect(request.url)
    
    return render_template('import_json.html')

@app.route('/import/json/map-fields')
@login_required
def map_json_fields():
    """JSON field mapping interface"""
    if 'json_analysis' not in session:
        flash('No JSON file data found. Please upload a file first.', 'error')
        return redirect(url_for('import_json'))
    
    analysis = session['json_analysis']
    filename = session.get('json_filename', 'uploaded_file.json')
    
    # Generate automatic field mapping
    importer = JSONImporter()
    suggested_mapping = importer.map_fields(analysis['fields'], analysis['data_type'])
    target_fields = importer.supported_formats[analysis['data_type']]
    
    return render_template('map_json_fields.html', 
                         analysis=analysis, 
                         filename=filename,
                         suggested_mapping=suggested_mapping,
                         target_fields=target_fields)

@app.route('/import/json/execute', methods=['POST'])
@login_required
def execute_json_import():
    """Execute JSON data import with user-defined field mapping"""
    if 'json_analysis' not in session:
        flash('No JSON file data found. Please start over.', 'error')
        return redirect(url_for('import_json'))
    
    analysis = session['json_analysis']
    data_type = analysis['data_type']
    
    # Load data from temporary file
    data_path = session.get('json_data_path')
    if not data_path or not os.path.exists(data_path):
        flash('Data file not found. Please upload the file again.', 'error')
        return redirect(url_for('import_json'))
    
    try:
        with open(data_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except Exception as e:
        flash(f'Error reading data file: {str(e)}', 'error')
        return redirect(url_for('import_json'))
    
    # Get field mapping from form
    field_mapping = {}
    for source_field in analysis['fields']:
        target_field = request.form.get(f'mapping_{source_field}')
        if target_field and target_field != 'skip':
            field_mapping[source_field] = target_field
    
    if not field_mapping:
        flash('Please map at least one field', 'error')
        return redirect(url_for('map_json_fields'))
    
    try:
        # Validate and transform data
        importer = JSONImporter()
        valid_records, validation_errors = importer.validate_data(data, data_type, field_mapping)
        
        if not valid_records:
            flash('No valid records found after validation', 'error')
            if validation_errors:
                for error in validation_errors[:5]:  # Show first 5 errors
                    flash(error, 'error')
            return redirect(url_for('map_json_fields'))
        
        # Import data
        import_result = importer.import_data(valid_records, data_type)
        
        # Clear session data and cleanup temp files
        session.pop('json_analysis', None)
        session.pop('json_filename', None)
        data_path = session.pop('json_data_path', None)
        if data_path and os.path.exists(data_path):
            try:
                os.unlink(data_path)
            except:
                pass  # Ignore cleanup errors
        
        # Show results
        if import_result['success']:
            flash(f'Successfully imported {import_result["imported"]} {data_type} records', 'success')
        else:
            flash(f'Import completed with errors. Imported {import_result["imported"]} of {import_result["total"]} records', 'warning')
            for error in import_result.get('errors', [])[:10]:  # Show first 10 errors
                flash(error, 'error')
        
        # Show validation errors if any
        if validation_errors:
            flash(f'{len(validation_errors)} records had validation errors and were skipped', 'warning')
        
        # Redirect based on data type
        if data_type == 'customers':
            return redirect(url_for('customers'))
        elif data_type == 'cylinders':
            return redirect(url_for('cylinders'))
        else:
            return redirect(url_for('index'))
            
    except Exception as e:
        flash(f'Import failed: {str(e)}', 'error')
        return redirect(url_for('map_json_fields'))

# Data Import routes
@app.route('/import')
@login_required
def import_data():
    """Data import dashboard with JSON and Access support"""
    return render_template('import_data.html', access_available=ACCESS_AVAILABLE)

@app.route('/import/from_replit', methods=['GET', 'POST'])
@admin_required
def import_from_replit():
    """Import data from another Replit Varasicyl database"""
    if request.method == 'GET':
        return render_template('import_from_replit.html')
    
    # Get connection details from form
    source_database_url = request.form.get('source_database_url', '').strip()
    data_types = request.form.getlist('data_types')  # ['customers', 'cylinders', 'rental_history']
    
    if not source_database_url:
        flash('Please provide a source database URL', 'error')
        return render_template('import_from_replit.html')
    
    if not data_types:
        flash('Please select at least one data type to import', 'error')
        return render_template('import_from_replit.html')
    
    try:
        from sqlalchemy import create_engine, text
        
        # Create connection to source database
        source_engine = create_engine(source_database_url)
        
        import_results = []
        total_imported = 0
        
        # Import each selected data type
        for data_type in data_types:
            try:
                with source_engine.connect() as conn:
                    # Get data from source database and convert to records
                    if data_type == 'customers':
                        query = text("SELECT * FROM customers")
                        result_proxy = conn.execute(query)
                        columns = result_proxy.keys()
                        records = [dict(zip(columns, row)) for row in result_proxy.fetchall()]
                        
                        # Import using JSON importer
                        from json_importer import JSONImporter
                        importer = JSONImporter()
                        valid_records, errors = importer.validate_data(records, 'customers', {})
                        result = importer.import_data(valid_records, 'customers')
                        
                    elif data_type == 'cylinders':
                        query = text("SELECT * FROM cylinders")
                        result_proxy = conn.execute(query)
                        columns = result_proxy.keys()
                        records = [dict(zip(columns, row)) for row in result_proxy.fetchall()]
                        
                        from json_importer import JSONImporter
                        importer = JSONImporter()
                        valid_records, errors = importer.validate_data(records, 'cylinders', {})
                        result = importer.import_data(valid_records, 'cylinders')
                        
                    elif data_type == 'rental_history':
                        query = text("SELECT * FROM rental_history")
                        result_proxy = conn.execute(query)
                        columns = result_proxy.keys()
                        records = [dict(zip(columns, row)) for row in result_proxy.fetchall()]
                        
                        from json_importer import JSONImporter
                        importer = JSONImporter()
                        valid_records, errors = importer.validate_data(records, 'rental_history', {})
                        result = importer.import_data(valid_records, 'rental_history')
                    
                    import_results.append({
                        'data_type': data_type,
                        'imported': result.get('imported', 0),
                        'errors': result.get('errors', []),
                        'success': result.get('success', False)
                    })
                    total_imported += result.get('imported', 0)
                    
            except Exception as e:
                import_results.append({
                    'data_type': data_type,
                    'imported': 0,
                    'errors': [f"Failed to import {data_type}: {str(e)}"],
                    'success': False
                })
        
        # Display results
        if total_imported > 0:
            flash(f'Successfully imported {total_imported} records from Replit database', 'success')
        
        for result in import_results:
            if result['errors']:
                for error in result['errors']:
                    flash(f"{result['data_type']}: {error}", 'warning')
        
        return render_template('import_from_replit.html', results=import_results)
        
    except Exception as e:
        flash(f'Error connecting to source database: {str(e)}', 'error')
        return render_template('import_from_replit.html')

@app.route('/import/rental-history', methods=['GET', 'POST'])
@admin_required
def import_rental_history():
    """Import rental history from JSON file"""
    if request.method == 'POST':
        try:
            # Check if file was uploaded
            if 'file' not in request.files:
                flash('No file selected', 'error')
                return redirect(url_for('import_rental_history'))
            
            file = request.files['file']
            if file.filename == '':
                flash('No file selected', 'error')
                return redirect(url_for('import_rental_history'))
            
            if not file.filename.lower().endswith('.json'):
                flash('Please select a JSON file', 'error')
                return redirect(url_for('import_rental_history'))
            
            # Read and parse JSON file
            file_content = file.read().decode('utf-8')
            rental_data = json.loads(file_content)
            
            # Validate JSON structure
            if not isinstance(rental_data, list):
                flash('JSON file must contain an array of rental history records', 'error')
                return redirect(url_for('import_rental_history'))
            
            # Import rental history records
            from db_service import RentalHistoryService
            imported_count = 0
            errors = []
            
            with RentalHistoryService() as rental_service:
                for i, record in enumerate(rental_data):
                    try:
                        # Create rental history record
                        history_record = {
                            'customer_id': record.get('customer_id', ''),
                            'customer_no': record.get('customer_no', ''),
                            'customer_name': record.get('customer_name', ''),
                            'customer_phone': record.get('customer_phone', ''),
                            'customer_email': record.get('customer_email', ''),
                            'customer_address': record.get('customer_address', ''),
                            'customer_city': record.get('customer_city', ''),
                            'customer_state': record.get('customer_state', ''),
                            'cylinder_id': record.get('cylinder_id', ''),
                            'cylinder_no': record.get('cylinder_no', ''),
                            'cylinder_custom_id': record.get('cylinder_custom_id', ''),
                            'cylinder_serial': record.get('cylinder_serial', ''),
                            'cylinder_type': record.get('cylinder_type', ''),
                            'cylinder_size': record.get('cylinder_size', ''),
                            'dispatch_date': record.get('dispatch_date', ''),
                            'return_date': record.get('return_date', ''),
                            'date_borrowed': record.get('date_borrowed', ''),
                            'date_returned': record.get('date_returned', ''),
                            'rental_days': record.get('rental_days', 0),
                            'location': record.get('location', ''),
                            'status': record.get('status', 'completed'),
                            'created_at': record.get('created_at', datetime.now().isoformat())
                        }
                        
                        # Create rental history record
                        rental_service.create_history_record(history_record)
                        imported_count += 1
                        
                    except Exception as e:
                        errors.append(f"Record {i+1}: {str(e)}")
                        continue
            
            # Show results
            if imported_count > 0:
                flash(f'Successfully imported {imported_count} rental history records', 'success')
            
            if errors:
                error_msg = f"Errors encountered: {len(errors)} records failed to import"
                if len(errors) <= 5:
                    error_msg += ": " + "; ".join(errors)
                flash(error_msg, 'warning')
            
            return redirect(url_for('rental_history'))
            
        except json.JSONDecodeError:
            flash('Invalid JSON file format', 'error')
        except Exception as e:
            flash(f'Error importing rental history: {str(e)}', 'error')
    
    return render_template('import_rental_history.html')

@app.route('/import/recreate-tables', methods=['POST'])
@admin_required
def recreate_tables():
    """Recreate database tables with exact schema matching"""
    try:
        from migrate_to_render import ReplitToRenderMigrator
        migrator = ReplitToRenderMigrator()
        
        # Recreate tables
        result = migrator.recreate_exact_tables()
        
        if 'error' in result:
            flash(f'Error recreating tables: {result["error"]}', 'error')
        else:
            flash(f'Tables recreated successfully: {", ".join(result["tables_created"])}', 'success')
            
            # Verify structure
            verification = migrator.verify_table_structure()
            if 'error' not in verification:
                flash('Table structure verified successfully', 'success')
        
        return redirect(url_for('import_data'))
        
    except Exception as e:
        flash(f'Error recreating tables: {str(e)}', 'error')
        return redirect(url_for('import_data'))

@app.route('/import/upload', methods=['POST'])
def upload_access_file():
    """Upload and connect to Access database"""
    if not ACCESS_AVAILABLE:
        flash('MS Access import functionality is not available', 'error')
        return redirect(url_for('index'))
    
    if 'access_file' not in request.files:
        flash('No file selected', 'error')
        return redirect(url_for('import_data'))
    
    file = request.files['access_file']
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('import_data'))
    
    if not file.filename.lower().endswith(('.mdb', '.accdb')):
        flash('Please select a valid Access database file (.mdb or .accdb)', 'error')
        return redirect(url_for('import_data'))
    
    try:
        # Save uploaded file temporarily with unique name to avoid conflicts
        import time
        timestamp = str(int(time.time()))
        temp_dir = tempfile.gettempdir()
        temp_path = os.path.join(temp_dir, f"temp_access_db{timestamp}.mdb")
        file.save(temp_path)
        
        # Try to connect
        importer = DataImporter()
        try:
            if importer.connect_to_access(temp_path):
                # Store file path in session
                session['access_file_path'] = temp_path
                session['access_file_name'] = file.filename
                
                # Get available tables
                tables = importer.get_available_tables()
                
                if tables:
                    flash(f'Successfully connected to {file.filename}. Found {len(tables)} tables.', 'success')
                    return render_template('select_tables.html', tables=tables, filename=file.filename)
                else:
                    flash('No tables found in the database', 'error')
                    return redirect(url_for('import_data'))
            else:
                flash('Failed to connect to Access database. Please check the file format and try again.', 'error')
                return redirect(url_for('import_data'))
        finally:
            # Always close connection to release file locks
            importer.close_connection()
            
    except Exception as e:
        flash(f'Error processing file: {str(e)}', 'error')
        if 'temp_path' in locals() and os.path.exists(temp_path):
            os.remove(temp_path)
        return redirect(url_for('import_data'))

@app.route('/import/preview/<table_name>')
def preview_table(table_name):
    """Preview table data and set up field mapping"""
    if 'access_file_path' not in session:
        flash('No Access file connected. Please upload a file first.', 'error')
        return redirect(url_for('import_data'))
    
    try:
        importer = DataImporter()
        if not importer.connect_to_access(session['access_file_path']):
            flash('Failed to reconnect to Access database', 'error')
            return redirect(url_for('import_data'))
        
        # Get table structure and preview data
        columns, preview_data = importer.preview_table(table_name)
        
        # Determine import type based on user selection
        import_type = request.args.get('type', 'customer')
        
        # Get suggested field mapping based on import type
        if import_type == 'transaction' or import_type == 'rental_history':
            suggested_mapping = importer.suggest_transaction_field_mapping(table_name)
        else:
            suggested_mapping = importer.suggest_field_mapping(table_name, import_type)
        
        importer.close_connection()
        
        return render_template('map_fields.html', 
                             table_name=table_name,
                             columns=columns,
                             preview_data=preview_data,
                             import_type=import_type,
                             suggested_mapping=suggested_mapping,
                             filename=session.get('access_file_name', 'Unknown'))
        
    except Exception as e:
        flash(f'Error previewing table: {str(e)}', 'error')
        return redirect(url_for('import_data'))

@app.route('/import/execute', methods=['POST'])
def execute_import():
    """Execute the data import"""
    if 'access_file_path' not in session:
        flash('No Access file connected. Please upload a file first.', 'error')
        return redirect(url_for('import_data'))
    
    try:
        table_name = request.form.get('table_name')
        import_type = request.form.get('import_type')
        skip_duplicates = request.form.get('skip_duplicates') == 'on'
        
        # Build field mapping from form data
        field_mapping = {}
        for key, value in request.form.items():
            if key.startswith('mapping_') and value:
                target_field = key.replace('mapping_', '')
                field_mapping[target_field] = value
        
        if not field_mapping:
            flash('Please map at least one field', 'error')
            return redirect(url_for('preview_table', table_name=table_name, type=import_type))
        
        # Execute instant import - no DataImporter needed
        from instant_importer import InstantImporter
        instant_importer = InstantImporter()
        
        print(f"🚀 Starting INSTANT {import_type.upper()} import...")
        imported, skipped, errors = instant_importer.instant_import(
            session['access_file_path'], 
            table_name, 
            field_mapping,
            import_type
        )
        
        if import_type == 'customer':
            item_type = 'customers'
        elif import_type == 'cylinder':
            item_type = 'cylinders'
        elif import_type == 'transaction':
            item_type = 'transactions'
        elif import_type == 'rental_history':
            item_type = 'rental history records'
        else:
            flash('Invalid import type', 'error')
            return redirect(url_for('import_data'))
        
        # Clean up temp file with retry logic for Windows
        temp_file_path = session.get('access_file_path')
        if temp_file_path and os.path.exists(temp_file_path):
            max_retries = 5
            for attempt in range(max_retries):
                try:
                    import time
                    time.sleep(0.2)  # Small delay to let file handles close
                    os.remove(temp_file_path)
                    print(f"Successfully removed temporary file: {temp_file_path}")
                    break
                except PermissionError as pe:
                    if attempt == max_retries - 1:
                        print(f"Warning: Could not remove temp file after {max_retries} attempts: {pe}")
                    else:
                        time.sleep(1)  # Wait longer between retries
                except Exception as e:
                    print(f"Warning: Error removing temp file: {e}")
                    break
        
        session.pop('access_file_path', None)
        session.pop('access_file_name', None)
        
        # Show results
        if imported > 0:
            flash(f'Successfully imported {imported} {item_type}', 'success')
        if skipped > 0:
            flash(f'Skipped {skipped} records (duplicates or missing data)', 'warning')
        if errors:
            for error in errors[:5]:  # Show first 5 errors
                flash(error, 'error')
            if len(errors) > 5:
                flash(f'... and {len(errors) - 5} more errors', 'error')
        
        # Show instant completion
        flash(f'🚀 INSTANT IMPORT COMPLETE! Processed {imported:,} {item_type} with zero overhead', 'success')
        
        # Redirect to appropriate page
        if import_type == 'customer':
            return redirect(url_for('customers'))
        elif import_type == 'rental_history':
            return redirect(url_for('rental_history'))
        else:
            return redirect(url_for('cylinders'))
        
    except Exception as e:
        flash(f'Error during import: {str(e)}', 'error')
        # Clean up on error
        temp_file_path = session.get('access_file_path')
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.remove(temp_file_path)
            except:
                pass
        session.pop('access_file_path', None)
        session.pop('access_file_name', None)
        return redirect(url_for('import_data'))

@app.route('/import/cancel')
def cancel_import():
    """Cancel import and clean up"""
    if 'access_file_path' in session:
        if os.path.exists(session['access_file_path']):
            os.remove(session['access_file_path'])
        session.pop('access_file_path', None)
        session.pop('access_file_name', None)
    
    flash('Import cancelled', 'info')
    return redirect(url_for('import_data'))

# Global Search route
@app.route('/search')
@login_required
def global_search():
    """Global search across customers and cylinders"""
    query = request.args.get('q', '').strip()
    
    results = {
        'customers': [],
        'cylinders': [],
        'query': query,
        'total_results': 0
    }
    
    if query:
        # Search customers
        with CustomerService() as customer_service:
            customers, _ = customer_service.get_all(page=1, per_page=10000)
        with CustomerService() as customer_service:
            customers, _ = customer_service.get_all(page=1, per_page=10000)
        results['customers'] = customer_results
        
        # Search cylinders
        with CylinderService() as cylinder_service:
            cylinder_results = cylinder_service.search(query)
        results['cylinders'] = cylinder_results
        
        results['total_results'] = len(customer_results) + len(cylinder_results)
    
    return render_template('search_results.html', **results)

@app.route('/users/delete/<user_id>', methods=['POST'])
@admin_required
def delete_user(user_id):
    """Delete user (admin only)"""
    try:
        if user_manager.delete_user(user_id):
            flash('User deleted successfully', 'success')
        else:
            flash('User not found', 'error')
    except ValueError as e:
        flash(str(e), 'error')
    except Exception as e:
        flash(f'Error deleting user: {str(e)}', 'error')
    
    return redirect(url_for('users'))

@app.route('/cylinders/rent/<cylinder_id>', methods=['POST'])
@user_or_admin_required
def rent_cylinder(cylinder_id):
    """Rent a cylinder to a customer"""
    customer_id = request.form.get('customer_id')
    rental_date = request.form.get('rental_date', '').strip()
    
    if not customer_id:
        flash('Please select a customer', 'error')
        return redirect(url_for('cylinders'))
    
    # Verify customer exists
    with CustomerService() as customer_service:
        customer = customer_service.get_by_id(customer_id)
    if not customer:
        flash('Customer not found', 'error')
        return redirect(url_for('cylinders'))
    
    # Convert rental_date to ISO format if provided
    rental_date_iso = None
    if rental_date:
        try:
            from datetime import datetime
            # Parse datetime-local format (YYYY-MM-DDTHH:MM) and convert to ISO
            dt = datetime.fromisoformat(rental_date)
            rental_date_iso = dt.isoformat()
        except ValueError:
            flash('Invalid rental date format', 'error')
            return redirect(url_for('cylinders'))
    
    # Get cylinder for display
    with CylinderService() as cylinder_service:
        cylinder = cylinder_service.get_by_id(cylinder_id)
        if cylinder:
            if hasattr(cylinder, 'custom_id'):
                display_id = cylinder.custom_id or cylinder.serial_number or cylinder.id
            else:
                display_id = cylinder.get('custom_id') or cylinder.get('serial_number') or cylinder.get('id', 'Unknown')
        else:
            display_id = 'Unknown'
        
        # Rent the cylinder with optional rental date
        success = cylinder_service.rent_cylinder(cylinder_id, customer_id, rental_date_iso)
    
    # Get customer name for flash message
    if hasattr(customer, 'customer_name'):
        customer_name = customer.customer_name or 'customer'
    elif isinstance(customer, dict):
        customer_name = customer.get('customer_name') or customer.get('name', 'customer')
    else:
        customer_name = 'customer'
    
    if success:
        flash(f'Cylinder {display_id} rented to {customer_name} successfully', 'success')
    else:
        flash(f'Error renting cylinder {display_id}', 'error')
    
    return redirect(url_for('cylinders'))

@app.route('/cylinders/return/<cylinder_id>', methods=['POST'])
@user_or_admin_required
def return_cylinder(cylinder_id):
    """Return a cylinder from rental"""
    return_date = request.form.get('return_date')
    with CylinderService() as cylinder_service:
        success = cylinder_service.return_cylinder(cylinder_id, return_date)
    
    if success:
        flash('Cylinder returned successfully', 'success')
    else:
        flash('Error returning cylinder', 'error')
    
    return redirect(url_for('cylinders'))

@app.route('/cylinder/<cylinder_id>/return/customer/<customer_id>', methods=['POST'])
@user_or_admin_required
def return_cylinder_custom(cylinder_id, customer_id):
    """Return cylinder with custom date from customer details page"""
    from datetime import datetime
    return_date = request.form.get('return_date', datetime.now().strftime('%Y-%m-%d'))
    
    with CylinderService() as cylinder_service:
        success = cylinder_service.return_cylinder(cylinder_id, return_date)
    
    if success:
        flash(f'Cylinder returned successfully on {return_date}', 'success')
    else:
        flash('Failed to return cylinder', 'error')
    
    return redirect(url_for('customer_details', customer_id=customer_id, tab='active'))

@app.route('/customers/<customer_id>/bulk_cylinders', methods=['GET', 'POST'])
@user_or_admin_required
def bulk_cylinder_management(customer_id):
    """Bulk cylinder rental/return management"""
    with CustomerService() as customer_service:
        customer_obj = customer_service.get_by_id(customer_id)
        
        if not customer_obj:
            flash('Customer not found', 'error')
            return redirect(url_for('customers'))
        
        # Convert customer to dictionary within session to avoid detached instance errors
        customer = {
            'id': customer_obj.id,
            'customer_no': customer_obj.customer_no,
            'customer_name': customer_obj.customer_name,
            'customer_email': customer_obj.customer_email,
            'customer_phone': customer_obj.customer_phone,
            'customer_address': customer_obj.customer_address,
            'customer_city': customer_obj.customer_city,
            'customer_state': customer_obj.customer_state
        }
    
    if request.method == 'GET':
        # Get current rentals for this customer
        with CylinderService() as cylinder_service:
            current_rentals = cylinder_service.get_by_customer(customer_id)
        return render_template('bulk_cylinder_management.html', 
                             customer=customer, 
                             current_rentals=current_rentals)
    
    cylinder_ids_text = request.form.get('cylinder_ids', '').strip()
    action = request.form.get('action', 'rent')
    date = request.form.get('date', '').strip()
    
    if not cylinder_ids_text:
        flash('Please enter at least one cylinder ID', 'error')
        return redirect(url_for('bulk_cylinder_management', customer_id=customer_id))
    
    if not date:
        flash('Please select a date', 'error')
        return redirect(url_for('bulk_cylinder_management', customer_id=customer_id))
    
    # Parse cylinder IDs from text (support both comma-separated and line-separated)
    cylinder_ids = []
    for line in cylinder_ids_text.replace(',', '\n').split('\n'):
        cylinder_id = line.strip()
        if cylinder_id:
            cylinder_ids.append(cylinder_id)
    
    if not cylinder_ids:
        flash('No valid cylinder IDs found', 'error')
        return redirect(url_for('bulk_cylinder_management', customer_id=customer_id))
    
    processed = 0
    skipped = 0
    errors = []
    
    for cylinder_id in cylinder_ids:
        with CylinderService() as cylinder_service:
            cylinder = cylinder_service.find_by_any_identifier(cylinder_id)
        
        if not cylinder:
            errors.append(f'"{cylinder_id}": Not found in database')
            skipped += 1
            continue
        
        # Use the actual system ID for operations
        actual_cylinder_id = cylinder.get('id')
        cylinder_display = cylinder.get('custom_id') or cylinder.get('serial_number') or actual_cylinder_id
        
        if action == 'rent':
            # Check if cylinder is available
            if cylinder.get('status', '').lower() != 'available':
                errors.append(f'"{cylinder_display}": Not available (current status: {cylinder.get("status", "unknown")})')
                skipped += 1
                continue
            
            # Rent the cylinder with custom date
            # Convert date from YYYY-MM-DD to datetime ISO format
            rental_datetime = f"{date}T00:00:00"
            with CylinderService() as cylinder_service:
                success = cylinder_service.rent_cylinder(actual_cylinder_id, customer_id, rental_datetime)
            if success:
                processed += 1
            else:
                errors.append(f'"{cylinder_display}": Failed to dispatch')
                skipped += 1
        
        elif action == 'return':
            # Debug info for troubleshooting
            cylinder_status = cylinder.get('status', '').lower()
            cylinder_rented_to = cylinder.get('rented_to')
            
            # Check if cylinder is rented/dispatched to this customer
            if cylinder_status not in ['rented', 'dispatched']:
                errors.append(f'"{cylinder_display}": Not rented/dispatched (current status: {cylinder.get("status", "unknown")})')
                skipped += 1
                continue
            elif cylinder_rented_to != customer_id:
                errors.append(f'"{cylinder_display}": Rented/dispatched to different customer (ID: {cylinder_rented_to})')
                skipped += 1
                continue
            
            # Return the cylinder with custom date
            # Convert date from YYYY-MM-DD to datetime ISO format
            return_datetime = f"{date}T00:00:00"
            with CylinderService() as cylinder_service:
                success = cylinder_service.return_cylinder(actual_cylinder_id, return_datetime)
            if success:
                processed += 1
            else:
                errors.append(f'"{cylinder_display}": Failed to return - database error')
                skipped += 1
    
    # Create summary message
    customer_name = customer.get('customer_name') or customer.get('name', 'Unknown Customer')
    if action == 'rent':
        flash(f'Successfully dispatched {processed} cylinders to {customer_name}', 'success')
    else:
        flash(f'Successfully returned {processed} cylinders from {customer_name}', 'success')
    
    if skipped > 0:
        flash(f'{skipped} cylinders were skipped due to errors', 'warning')
        
    # Show detailed errors if any
    if errors:
        error_msg = 'Details: ' + '; '.join(errors[:5])  # Show first 5 errors
        if len(errors) > 5:
            error_msg += f' and {len(errors) - 5} more...'
        flash(error_msg, 'info')
    
    return redirect(url_for('bulk_cylinder_management', customer_id=customer_id))

@app.route('/api/customer/<customer_id>/rentals')
@login_required
def get_customer_rentals(customer_id):
    """API endpoint to get current rentals for a customer"""
    with CylinderService() as cylinder_service:
        rentals = cylinder_service.get_by_customer(customer_id)
    
    # Add rental days calculation
    for rental in rentals:
        if 'date_borrowed' in rental and rental['date_borrowed']:
            from datetime import datetime
            try:
                if isinstance(rental['date_borrowed'], str):
                    borrowed_date = datetime.fromisoformat(rental['date_borrowed'].replace('Z', '+00:00'))
                else:
                    borrowed_date = rental['date_borrowed']
                rental['rental_days'] = (datetime.now().replace(tzinfo=borrowed_date.tzinfo) - borrowed_date).days
            except:
                rental['rental_days'] = 0
        else:
            rental['rental_days'] = 0
    
    return jsonify({'rentals': rentals})

@app.route('/archive_data', methods=['POST'])
@login_required
@admin_required
def archive_data():
    """Archive old data (admin only)"""
    try:
        months_old = int(request.form.get('months', 6))
        if months_old < 1:
            months_old = 6
        
        # Archive both cylinder and customer data
        with CylinderService() as cylinder_service:
            cylinder_result = cylinder_service.archive_old_data(months_old)
        with CustomerService() as customer_service:
            customers, _ = customer_service.get_all(page=1, per_page=10000)
        with CustomerService() as customer_service:
            customers, _ = customer_service.get_all(page=1, per_page=10000)
        
        # Combine results
        total_archived = cylinder_result.get('archived_count', 0) + customer_result.get('archived_count', 0)
        total_remaining = cylinder_result.get('remaining_count', 0) + customer_result.get('remaining_count', 0)
        
        # Check for errors
        if 'error' in cylinder_result or 'error' in customer_result:
            errors = []
            if 'error' in cylinder_result:
                errors.append(f"Cylinders: {cylinder_result['error']}")
            if 'error' in customer_result:
                errors.append(f"Customers: {customer_result['error']}")
            flash(f'Archive failed: {"; ".join(errors)}', 'error')
        elif total_archived > 0:
            archive_files = []
            if cylinder_result.get('archived_count', 0) > 0:
                archive_files.append(cylinder_result.get('archive_file', ''))
            if customer_result.get('archived_count', 0) > 0:
                archive_files.append(customer_result.get('archive_file', ''))
            
            flash(f'Successfully archived {total_archived} old records ({cylinder_result.get("archived_count", 0)} cylinders, {customer_result.get("archived_count", 0)} customers). Archives saved.', 'success')
        else:
            flash('No old data found to archive', 'info')
        

        
    except ValueError:
        flash('Invalid months value provided', 'error')
    except Exception as e:
        flash(f'Error during archiving: {str(e)}', 'error')
    
    return redirect(url_for('cylinders'))

@app.route('/bulk_rental_management')
@login_required
def bulk_rental_management():
    """Dedicated page for bulk cylinder rental management"""
    # Get customers and convert to dict within service context
    with CustomerService() as customer_service:
        customers, _ = customer_service.get_all(page=1, per_page=500)  # Get all customers
        
        # Convert customers to dict format within active session
        customers_dict = []
        for customer in customers:
            if isinstance(customer, dict):
                customers_dict.append(customer)
            else:
                # Convert SQLAlchemy object to dict within active session
                try:
                    customer_dict = {
                        'id': customer.id if hasattr(customer, 'id') else '',
                        'customer_no': customer.customer_no if hasattr(customer, 'customer_no') and customer.customer_no else '',
                        'customer_name': customer.customer_name if hasattr(customer, 'customer_name') and customer.customer_name else '',
                        'customer_email': customer.customer_email if hasattr(customer, 'customer_email') and customer.customer_email else '',
                        'customer_phone': customer.customer_phone if hasattr(customer, 'customer_phone') and customer.customer_phone else '',
                        'customer_address': customer.customer_address if hasattr(customer, 'customer_address') and customer.customer_address else '',
                        'customer_city': customer.customer_city if hasattr(customer, 'customer_city') and customer.customer_city else '',
                        'customer_state': customer.customer_state if hasattr(customer, 'customer_state') and customer.customer_state else ''
                    }
                    customers_dict.append(customer_dict)
                except Exception as e:
                    app.logger.error(f"Error converting customer: {str(e)}")
                    continue
    
    # Get cylinders and convert to dict within service context
    with CylinderService() as cylinder_service:
        cylinders, _ = cylinder_service.get_all()
        
        # Convert cylinders to dict format within active session
        cylinders_dict = []
        for cylinder in cylinders:
            if isinstance(cylinder, dict):
                cylinder_dict = cylinder
            else:
                # Convert SQLAlchemy object to dict within active session
                try:
                    from datetime import datetime
                    cylinder_dict = {
                        'id': cylinder.id if hasattr(cylinder, 'id') else '',
                        'custom_id': cylinder.custom_id if hasattr(cylinder, 'custom_id') and cylinder.custom_id else '',
                        'serial_number': cylinder.serial_number if hasattr(cylinder, 'serial_number') and cylinder.serial_number else '',
                        'display_id': (cylinder.custom_id if hasattr(cylinder, 'custom_id') and cylinder.custom_id else 
                                     cylinder.serial_number if hasattr(cylinder, 'serial_number') and cylinder.serial_number else 
                                     f"ID-{cylinder.id[:8]}" if hasattr(cylinder, 'id') else 'Unknown'),
                        'type': cylinder.type if hasattr(cylinder, 'type') and cylinder.type else '',
                        'size': cylinder.size if hasattr(cylinder, 'size') and cylinder.size else '',
                        'status': cylinder.status if hasattr(cylinder, 'status') and cylinder.status else '',
                        'location': cylinder.location if hasattr(cylinder, 'location') and cylinder.location else '',
                        'rented_to': cylinder.rented_to if hasattr(cylinder, 'rented_to') and cylinder.rented_to else '',
                        'customer_name': cylinder.customer_name if hasattr(cylinder, 'customer_name') and cylinder.customer_name else '',
                        'rental_days': (datetime.utcnow() - cylinder.date_borrowed).days if hasattr(cylinder, 'date_borrowed') and cylinder.date_borrowed else 0,
                        'date_borrowed': cylinder.date_borrowed.isoformat() if hasattr(cylinder, 'date_borrowed') and cylinder.date_borrowed else ''
                    }
                    
                    # Add customer name for rented cylinders (use existing customer_name if available)
                    if cylinder_dict.get('rented_to') and not cylinder_dict.get('customer_name'):
                        cylinder_dict['customer_name'] = 'Rented Customer'
                    
                    cylinders_dict.append(cylinder_dict)
                except Exception as e:
                    app.logger.error(f"Error converting cylinder: {str(e)}")
                    continue
    
    return render_template('bulk_rental_management.html', customers=customers_dict, cylinders=cylinders_dict)

@app.route('/bulk_rental_management/process', methods=['POST'])
@login_required
def process_bulk_rental():
    """Process bulk cylinder rental/return operations"""
    customer_id = request.form.get('customer_id', '').strip()
    action = request.form.get('action', 'rent')
    date = request.form.get('date', '').strip()
    cylinder_ids_text = request.form.get('cylinder_ids', '').strip()
    
    if not customer_id:
        flash('Please select a customer', 'error')
        return redirect(url_for('bulk_rental_management'))
    
    with CustomerService() as customer_service:
        customer = customer_service.get_by_id(customer_id)
    if not customer:
        flash('Customer not found', 'error')
        return redirect(url_for('bulk_rental_management'))
    
    if not cylinder_ids_text:
        flash('Please enter at least one cylinder ID', 'error')
        return redirect(url_for('bulk_rental_management'))
    
    if not date:
        flash('Please select a date', 'error')
        return redirect(url_for('bulk_rental_management'))
    
    # Parse cylinder IDs from text (support both comma-separated and line-separated)
    cylinder_ids = []
    for line in cylinder_ids_text.replace(',', '\n').split('\n'):
        cylinder_id = line.strip()
        if cylinder_id:
            cylinder_ids.append(cylinder_id)
    
    if not cylinder_ids:
        flash('No valid cylinder IDs found', 'error')
        return redirect(url_for('bulk_rental_management'))
    
    processed = 0
    skipped = 0
    errors = []
    success_cylinders = []
    
    for cylinder_id in cylinder_ids:
        with CylinderService() as cylinder_service:
            cylinder = cylinder_service.find_by_any_identifier(cylinder_id)
        
        if not cylinder:
            errors.append(f'Cylinder {cylinder_id}: Not found in database')
            skipped += 1
            continue
        
        # Use the actual system ID for operations
        actual_cylinder_id = cylinder.get('id')
        cylinder_display = cylinder.get('custom_id') or cylinder.get('serial_number') or actual_cylinder_id
        
        if action == 'rent':
            # Check if cylinder is available
            if cylinder.get('status', '').lower() != 'available':
                errors.append(f'"{cylinder_display}": Not available (current status: {cylinder.get("status", "unknown")})')
                skipped += 1
                continue
            
            # Rent the cylinder with custom date
            # Convert date from YYYY-MM-DD to datetime ISO format
            rental_datetime = f"{date}T00:00:00"
            with CylinderService() as cylinder_service:
                success = cylinder_service.rent_cylinder(actual_cylinder_id, customer_id, rental_datetime)
            if success:
                processed += 1
                success_cylinders.append(cylinder_display)
            else:
                errors.append(f'"{cylinder_display}": Failed to dispatch')
                skipped += 1
        
        elif action == 'return':
            # Check if cylinder is rented/dispatched to this customer
            cylinder_status = cylinder.get('status', '').lower()
            if cylinder_status not in ['rented', 'dispatched'] or cylinder.get('rented_to') != customer_id:
                errors.append(f'"{cylinder_display}": Not rented/dispatched to this customer (status: {cylinder.get("status", "unknown")})')
                skipped += 1
                continue
            
            # Return the cylinder with custom date
            # Convert date from YYYY-MM-DD to datetime ISO format  
            return_datetime = f"{date}T00:00:00"
            with CylinderService() as cylinder_service:
                success = cylinder_service.return_cylinder(actual_cylinder_id, return_datetime)
            if success:
                processed += 1
                success_cylinders.append(cylinder_display)
            else:
                errors.append(f'"{cylinder_display}": Failed to return')
                skipped += 1
    
    # Create summary message - use safe dict access since customer is already converted to dict
    if isinstance(customer, dict):
        customer_name = customer.get('customer_name') or customer.get('name', 'Unknown Customer')
    else:
        customer_name = 'Unknown Customer'
        
    if action == 'rent':
        if processed > 0:
            flash(f'Successfully dispatched {processed} cylinders ({", ".join(success_cylinders[:5])}{", ..." if len(success_cylinders) > 5 else ""}) to {customer_name}', 'success')
    else:
        if processed > 0:
            flash(f'Successfully returned {processed} cylinders ({", ".join(success_cylinders[:5])}{", ..." if len(success_cylinders) > 5 else ""}) from {customer_name}', 'success')
    
    if skipped > 0:
        flash(f'{skipped} cylinders were skipped due to errors', 'warning')
        
    # Show detailed errors if any
    if errors:
        error_msg = 'Details: ' + '; '.join(errors[:5])  # Show first 5 errors
        if len(errors) > 5:
            error_msg += f' and {len(errors) - 5} more...'
        flash(error_msg, 'info')
    
    return redirect(url_for('bulk_rental_management'))

@app.route('/customers/<customer_id>/active_dispatches')
@login_required
def customer_active_dispatches(customer_id):
    """View active dispatches for a specific customer with pagination"""
    # Get customer details
    with CustomerService() as customer_service:
        customers, _ = customer_service.get_all(page=1, per_page=10000)
    with CustomerService() as customer_service:
        customers, _ = customer_service.get_all(page=1, per_page=10000)
    if not customer:
        flash('Customer not found', 'error')
        return redirect(url_for('customers'))
    
    # Get pagination parameters
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 25))
    
    # Get all cylinders rented to this customer using the service
    from db_service import CylinderService
    
    with CylinderService() as cylinder_service:
        customer_cylinders = cylinder_service.get_by_customer(customer_id)
    
    # Convert to dictionaries and add rental info - data already includes rental_days from service
    customer_cylinders_dict = []
    for cylinder in customer_cylinders:
        if isinstance(cylinder, dict):
            customer_cylinders_dict.append(cylinder)
        else:
            # Convert SQLAlchemy object to dict
            cylinder_dict = {
                'id': cylinder.id,
                'custom_id': cylinder.custom_id or '',
                'serial_number': cylinder.serial_number or '',
                'display_id': cylinder.custom_id or cylinder.serial_number or f"ID-{cylinder.id[:8]}",
                'type': cylinder.type or '',
                'size': cylinder.size or '',
                'status': cylinder.status or '',
                'location': cylinder.location or '',
                'rental_days': (datetime.utcnow() - cylinder.date_borrowed).days if cylinder.date_borrowed else 0,
                'rental_months': ((datetime.utcnow() - cylinder.date_borrowed).days // 30) if cylinder.date_borrowed else 0,
                'date_borrowed': cylinder.date_borrowed.isoformat() if cylinder.date_borrowed else ''
            }
            customer_cylinders_dict.append(cylinder_dict)
    
    customer_cylinders = customer_cylinders_dict
    
    # Sort by rental days (longest first)
    customer_cylinders.sort(key=lambda x: x.get('rental_days', 0), reverse=True)
    
    # Pagination
    total_cylinders = len(customer_cylinders)
    start = (page - 1) * per_page
    end = start + per_page
    cylinders_paginated = customer_cylinders[start:end]
    
    # Calculate pagination info
    total_pages = (total_cylinders + per_page - 1) // per_page
    has_prev = page > 1
    has_next = page < total_pages
    
    pagination_info = {
        'page': page,
        'per_page': per_page,
        'total': total_cylinders,
        'total_pages': total_pages,
        'has_prev': has_prev,
        'has_next': has_next,
        'prev_num': page - 1 if has_prev else None,
        'next_num': page + 1 if has_next else None,
        'start_index': start + 1 if cylinders_paginated else 0,
        'end_index': min(end, total_cylinders)
    }
    
    return render_template('customer_active_dispatches.html', 
                          customer=customer, 
                          customer_cylinders=cylinders_paginated,
                          pagination=pagination_info)

# Reports routes
@app.route('/reports')
@login_required
def reports():
    """Data reports and export page"""
    # Use PostgreSQL services for reports
    with CustomerService() as customer_service:
        customers, _ = customer_service.get_all(page=1, per_page=10000)
    
    with CylinderService() as cylinder_service:
        cylinders, _ = cylinder_service.get_all(page=1, per_page=20000)
    
    # PostgreSQL services already return dictionaries
    cylinders_dict = cylinders
    customers_dict = []
    
    for customer in customers:
        customer_dict = customer if isinstance(customer, dict) else {
            'id': customer.id,
            'customer_name': customer.customer_name or '',
            'customer_no': customer.customer_no or ''
        }
        
        # Add rental count for sorting customers
        rented_cylinders = [c for c in cylinders_dict if c.get('rented_to') == customer_dict['id']]
        customer_dict['rental_count'] = len(rented_cylinders)
        customers_dict.append(customer_dict)
    
    # Calculate stats
    active_rentals = len([c for c in cylinders_dict if c.get('status', '').lower() == 'rented'])
    
    # Sort customers by rental count descending
    customers_dict.sort(key=lambda x: x.get('rental_count', 0), reverse=True)
    
    stats = {
        'total_customers': len(customers_dict),
        'total_cylinders': len(cylinders_dict),
        'active_rentals': active_rentals
    }
    
    return render_template('reports.html', stats=stats, customers=customers_dict)

@app.route('/export/customers.csv')
@login_required
def export_customers_csv():
    """Export all customers to CSV"""
    try:
        with CustomerService() as customer_service:
            customers, _ = customer_service.get_all(page=1, per_page=10000)
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        writer.writerow(['ID', 'Customer No', 'Name', 'Email', 'Phone', 'Address', 'City', 'State', 'APGST', 'CST', 'Created At', 'Updated At', 'Notes'])
        
        # Write customer data
        for customer in customers:
            # Safe attribute access for SQLAlchemy objects
            def safe_get(obj, attr, default=''):
                if hasattr(obj, attr):
                    return getattr(obj, attr) or default
                return obj.get(attr, default) if isinstance(obj, dict) else default
            
            writer.writerow([
                safe_get(customer, 'id'),
                safe_get(customer, 'customer_no'),
                safe_get(customer, 'customer_name') or safe_get(customer, 'name'),
                safe_get(customer, 'customer_email') or safe_get(customer, 'email'),
                safe_get(customer, 'customer_phone') or safe_get(customer, 'phone'),
                safe_get(customer, 'customer_address') or safe_get(customer, 'address'),
                safe_get(customer, 'customer_city'),
                safe_get(customer, 'customer_state'),
                safe_get(customer, 'customer_apgst'),
                safe_get(customer, 'customer_cst'),
                safe_get(customer, 'created_at'),
                safe_get(customer, 'updated_at'),
                safe_get(customer, 'notes')
            ])
        
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=customers_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'}
        )
    except Exception as e:
        flash(f'Error exporting customers: {str(e)}', 'error')
        return redirect(url_for('reports'))

@app.route('/export/cylinders.csv')
@login_required
def export_cylinders_csv():
    """Export all cylinders to CSV"""
    try:
        with CylinderService() as cylinder_service:
            cylinders, _ = cylinder_service.get_all(page=1, per_page=10000)
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        writer.writerow(['ID', 'Serial Number', 'Type', 'Size', 'Status', 'Location', 
                        'Pressure', 'Last Inspection', 'Next Inspection', 'Customer Name',
                        'Date Borrowed', 'Date Returned', 'Notes'])
        
        # Write cylinder data
        for cylinder in cylinders:
            # Safe attribute access for SQLAlchemy objects
            def safe_get(obj, attr, default=''):
                if hasattr(obj, attr):
                    return getattr(obj, attr) or default
                return obj.get(attr, default) if isinstance(obj, dict) else default
            
            # Get display ID
            display_id = safe_get(cylinder, 'custom_id') or safe_get(cylinder, 'serial_number') or safe_get(cylinder, 'id')
            
            # Format dispatch and return dates properly
            dispatch_date = safe_get(cylinder, 'date_borrowed') or safe_get(cylinder, 'rental_date')
            if dispatch_date and len(str(dispatch_date)) >= 10:
                dispatch_date = str(dispatch_date)[:10]  # Extract YYYY-MM-DD part
            
            return_date = safe_get(cylinder, 'date_returned')
            if return_date and len(str(return_date)) >= 10:
                return_date = str(return_date)[:10]  # Extract YYYY-MM-DD part
            
            writer.writerow([
                display_id,
                safe_get(cylinder, 'serial_number'),
                safe_get(cylinder, 'type'),
                safe_get(cylinder, 'size'),
                safe_get(cylinder, 'status'),
                safe_get(cylinder, 'location'),
                safe_get(cylinder, 'pressure'),
                safe_get(cylinder, 'last_inspection'),
                safe_get(cylinder, 'next_inspection'),
                safe_get(cylinder, 'customer_name'),
                dispatch_date,
                return_date,
                safe_get(cylinder, 'notes')
            ])
        
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=cylinders_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'}
        )
    except Exception as e:
        flash(f'Error exporting cylinders: {str(e)}', 'error')
        return redirect(url_for('reports'))

@app.route('/export/rental-activities.csv')
@login_required
def export_rental_activities_csv():
    """Export rental activities to CSV"""
    try:
        with CylinderService() as cylinder_service:
            cylinders, _ = cylinder_service.get_all(page=1, per_page=10000)
        with CustomerService() as customer_service:
            customers, _ = customer_service.get_all(page=1, per_page=10000)
        
        # Create customer lookup with safe access
        customer_lookup = {}
        for c in customers:
            if hasattr(c, 'id'):
                customer_lookup[c.id] = c
            elif isinstance(c, dict):
                customer_lookup[c.get('id')] = c
    
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        writer.writerow(['Cylinder ID', 'Serial Number', 'Type', 
                        'Customer Name', 'Customer Email', 'Date Borrowed', 'Date Returned', 
                        'Status', 'Rental Days'])
        
        # Write rental data
        for cylinder in cylinders:
            # Safe attribute access for SQLAlchemy objects
            def safe_get(obj, attr, default=''):
                if hasattr(obj, attr):
                    return getattr(obj, attr) or default
                return obj.get(attr, default) if isinstance(obj, dict) else default
            
            if safe_get(cylinder, 'rented_to') or safe_get(cylinder, 'date_borrowed'):
                customer_id = safe_get(cylinder, 'rented_to')
                customer = customer_lookup.get(customer_id, {})
                
                # Calculate rental days safely
                rental_days = ''
                dispatch_date_str = safe_get(cylinder, 'date_borrowed') or safe_get(cylinder, 'rental_date')
                if dispatch_date_str:
                    try:
                        if isinstance(dispatch_date_str, str):
                            dispatch_date = datetime.fromisoformat(dispatch_date_str.replace('Z', '+00:00'))
                        else:
                            dispatch_date = dispatch_date_str
                        return_date_str = safe_get(cylinder, 'date_returned')
                        if return_date_str:
                            if isinstance(return_date_str, str):
                                return_date = datetime.fromisoformat(return_date_str.replace('Z', '+00:00'))
                            else:
                                return_date = return_date_str
                            rental_days = (return_date - dispatch_date).days
                        else:
                            rental_days = (datetime.now() - dispatch_date).days
                    except:
                        rental_days = ''
                
                # Get display ID
                display_id = safe_get(cylinder, 'custom_id') or safe_get(cylinder, 'serial_number') or safe_get(cylinder, 'id')
                
                # Format dispatch and return dates properly
                dispatch_date = safe_get(cylinder, 'date_borrowed') or safe_get(cylinder, 'rental_date')
                if dispatch_date and len(str(dispatch_date)) >= 10:
                    dispatch_date = str(dispatch_date)[:10]  # Extract YYYY-MM-DD part
                
                return_date = safe_get(cylinder, 'date_returned')
                if return_date and len(str(return_date)) >= 10:
                    return_date = str(return_date)[:10]  # Extract YYYY-MM-DD part
                
                # Safe customer access
                def safe_customer_get(obj, attr, default=''):
                    if hasattr(obj, attr):
                        return getattr(obj, attr) or default
                    return obj.get(attr, default) if isinstance(obj, dict) else default
                
                writer.writerow([
                    display_id,
                    safe_get(cylinder, 'serial_number'),
                    safe_get(cylinder, 'type'),
                    safe_customer_get(customer, 'customer_name') or safe_customer_get(customer, 'name'),
                    safe_customer_get(customer, 'customer_email') or safe_customer_get(customer, 'email'),
                    dispatch_date,
                    return_date,
                    safe_get(cylinder, 'status'),
                    rental_days
                ])
        
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=rental_activities_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'}
        )
    except Exception as e:
        flash(f'Error exporting rental activities: {str(e)}', 'error')
        return redirect(url_for('reports'))

@app.route('/export/complete-data.csv')
@login_required
def export_complete_data_csv():
    """Export complete database to CSV"""
    try:
        with CustomerService() as customer_service:
            customers, _ = customer_service.get_all(page=1, per_page=10000)
        with CylinderService() as cylinder_service:
            cylinders, _ = cylinder_service.get_all(page=1, per_page=10000)
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write a complete report with all data
        writer.writerow(['=== COMPLETE DATABASE EXPORT ==='])
        writer.writerow(['Export Date:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        writer.writerow(['Total Customers:', len(customers)])
        writer.writerow(['Total Cylinders:', len(cylinders)])
        writer.writerow([])
        
        # Customers section
        writer.writerow(['=== CUSTOMERS ==='])
        writer.writerow(['ID', 'Customer No', 'Name', 'Email', 'Phone', 'Address', 'City', 'State', 'APGST', 'CST', 'Created At', 'Notes'])
        for customer in customers:
            # Safe attribute access for SQLAlchemy objects
            def safe_get(obj, attr, default=''):
                if hasattr(obj, attr):
                    return getattr(obj, attr) or default
                return obj.get(attr, default) if isinstance(obj, dict) else default
            
            writer.writerow([
                safe_get(customer, 'id'),
                safe_get(customer, 'customer_no'),
                safe_get(customer, 'customer_name') or safe_get(customer, 'name'),
                safe_get(customer, 'customer_email') or safe_get(customer, 'email'),
                safe_get(customer, 'customer_phone') or safe_get(customer, 'phone'),
                safe_get(customer, 'customer_address') or safe_get(customer, 'address'),
                safe_get(customer, 'customer_city'),
                safe_get(customer, 'customer_state'),
                safe_get(customer, 'customer_apgst'),
                safe_get(customer, 'customer_cst'),
                safe_get(customer, 'created_at'),
                safe_get(customer, 'notes')
            ])
        
        writer.writerow([])
        
        # Cylinders section
        writer.writerow(['=== CYLINDERS ==='])
        writer.writerow(['ID', 'Serial Number', 'Type', 'Size', 'Status', 'Location', 
                        'Pressure', 'Customer Name', 'Date Borrowed', 'Rental Days'])
        for cylinder in cylinders:
            # Safe attribute access for SQLAlchemy objects
            def safe_get_cyl(obj, attr, default=''):
                if hasattr(obj, attr):
                    return getattr(obj, attr) or default
                return obj.get(attr, default) if isinstance(obj, dict) else default
            
            # Calculate rental days safely
            rental_days = ''
            dispatch_date_str = safe_get_cyl(cylinder, 'date_borrowed') or safe_get_cyl(cylinder, 'rental_date')
            if dispatch_date_str:
                try:
                    if isinstance(dispatch_date_str, str):
                        dispatch_date = datetime.fromisoformat(dispatch_date_str.replace('Z', '+00:00'))
                    else:
                        dispatch_date = dispatch_date_str
                    return_date_str = safe_get_cyl(cylinder, 'date_returned')
                    if return_date_str:
                        if isinstance(return_date_str, str):
                            return_date = datetime.fromisoformat(return_date_str.replace('Z', '+00:00'))
                        else:
                            return_date = return_date_str
                        rental_days = (return_date - dispatch_date).days
                    else:
                        rental_days = (datetime.now() - dispatch_date).days
                except:
                    rental_days = ''
            
            # Get display ID
            display_id = safe_get_cyl(cylinder, 'custom_id') or safe_get_cyl(cylinder, 'serial_number') or safe_get_cyl(cylinder, 'id')
            
            # Format dispatch date properly
            dispatch_date = safe_get_cyl(cylinder, 'date_borrowed') or safe_get_cyl(cylinder, 'rental_date')
            if dispatch_date and len(str(dispatch_date)) >= 10:
                dispatch_date = str(dispatch_date)[:10]  # Extract YYYY-MM-DD part
            
            writer.writerow([
                display_id,
                safe_get_cyl(cylinder, 'serial_number'),
                safe_get_cyl(cylinder, 'type'),
                safe_get_cyl(cylinder, 'size'),
                safe_get_cyl(cylinder, 'status'),
                safe_get_cyl(cylinder, 'location'),
                safe_get_cyl(cylinder, 'pressure'),
                safe_get_cyl(cylinder, 'customer_name'),
                dispatch_date,
                rental_days
            ])
        
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=complete_database_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'}
        )
    except Exception as e:
        flash(f'Error exporting complete data: {str(e)}', 'error')
        return redirect(url_for('reports'))

@app.route('/export/customer-report', methods=['POST'])
@login_required
def export_customer_report():
    """Export individual customer report with dispatched cylinders sorted by rental days"""
    customer_id = request.form.get('customer_id')
    export_format = request.form.get('export_format', 'csv')
    
    if not customer_id:
        flash('Please select a customer', 'error')
        return redirect(url_for('reports'))
    
    try:
        # Get customer details
        with CustomerService() as customer_service:
            customers, _ = customer_service.get_all(page=1, per_page=10000)
        with CustomerService() as customer_service:
            customers, _ = customer_service.get_all(page=1, per_page=10000)
        if not customer:
            flash('Customer not found', 'error')
            return redirect(url_for('reports'))
        
        # Get all cylinders dispatched to this customer
        with CylinderService() as cylinder_service:
            all_cylinders, _ = cylinder_service.get_all(page=1, per_page=10000)
        
        # Safe access for customer ID
        def safe_get(obj, attr, default=''):
            if hasattr(obj, attr):
                return getattr(obj, attr) or default
            return obj.get(attr, default) if isinstance(obj, dict) else default
        
        customer_id_val = safe_get(customer, 'id')
        customer_cylinders = [c for c in all_cylinders if safe_get(c, 'rented_to') == customer_id_val]
        
        # Add rental days and sort by descending rental days
        for cylinder in customer_cylinders:
            try:
                date_borrowed = safe_get(cylinder, 'date_borrowed')
                if date_borrowed:
                    if isinstance(date_borrowed, str):
                        borrow_date = datetime.fromisoformat(date_borrowed.replace('Z', '+00:00'))
                    else:
                        borrow_date = date_borrowed
                    rental_days = (datetime.now() - borrow_date).days
                    # For dictionary-like access
                    if isinstance(cylinder, dict):
                        cylinder['rental_days'] = rental_days
                    else:
                        # For SQLAlchemy objects, we can't set attributes dynamically
                        setattr(cylinder, 'rental_days', rental_days)
                else:
                    if isinstance(cylinder, dict):
                        cylinder['rental_days'] = 0
                    else:
                        setattr(cylinder, 'rental_days', 0)
            except:
                if isinstance(cylinder, dict):
                    cylinder['rental_days'] = 0
                else:
                    setattr(cylinder, 'rental_days', 0)
        
        # Sort by rental days descending (longest rentals first)
        customer_cylinders.sort(key=lambda x: getattr(x, 'rental_days', 0) if hasattr(x, 'rental_days') else x.get('rental_days', 0), reverse=True)
        
        customer_name = safe_get(customer, 'customer_name') or safe_get(customer, 'name') or 'Unknown Customer'
        safe_filename = customer_name.replace(' ', '_').replace('/', '_')
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if export_format == 'pdf':
            return export_customer_pdf(customer, customer_cylinders, safe_filename, timestamp)
        else:  # Default to CSV
            return export_customer_csv(customer, customer_cylinders, safe_filename, timestamp)
    except Exception as e:
        flash(f'Error exporting customer report: {str(e)}', 'error')
        return redirect(url_for('reports'))

def export_customer_csv(customer, customer_cylinders, safe_filename, timestamp):
    """Export customer report as CSV"""
    try:
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Safe attribute access for SQLAlchemy objects
        def safe_get(obj, attr, default=''):
            if hasattr(obj, attr):
                return getattr(obj, attr) or default
            return obj.get(attr, default) if isinstance(obj, dict) else default
        
        customer_name = safe_get(customer, 'customer_name') or safe_get(customer, 'name') or 'Unknown Customer'
        
        # Customer details header
        writer.writerow([f'=== CUSTOMER REPORT: {customer_name} ==='])
        writer.writerow(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        writer.writerow([])
        
        # Customer information
        writer.writerow(['=== CUSTOMER DETAILS ==='])
        writer.writerow(['Customer No:', safe_get(customer, 'customer_no')])
        writer.writerow(['Name:', customer_name])
        writer.writerow(['Phone:', safe_get(customer, 'customer_phone') or safe_get(customer, 'phone')])
        writer.writerow(['Email:', safe_get(customer, 'customer_email') or safe_get(customer, 'email')])
        writer.writerow(['Address:', safe_get(customer, 'customer_address') or safe_get(customer, 'address')])
        writer.writerow(['City:', safe_get(customer, 'customer_city')])
        writer.writerow(['State:', safe_get(customer, 'customer_state')])
        writer.writerow(['Total Dispatched Cylinders:', len(customer_cylinders)])
        writer.writerow([])
        
        # Dispatched cylinders sorted by rental days
        writer.writerow(['=== DISPATCHED CYLINDERS (Sorted by Days Dispatched - Longest First) ==='])
        writer.writerow(['ID', 'Serial Number', 'Type', 'Size', 'Status', 
                        'Date Dispatched', 'Days Dispatched', 'Location', 'Pressure'])
        
        for cylinder in customer_cylinders:
            # Get display ID
            display_id = safe_get(cylinder, 'custom_id') or safe_get(cylinder, 'serial_number') or safe_get(cylinder, 'id')
            
            # Format dispatch date properly
            dispatch_date = safe_get(cylinder, 'date_borrowed') or safe_get(cylinder, 'rental_date')
            if dispatch_date and len(str(dispatch_date)) >= 10:
                dispatch_date = str(dispatch_date)[:10]  # Extract YYYY-MM-DD part
            
            rental_days = getattr(cylinder, 'rental_days', 0) if hasattr(cylinder, 'rental_days') else (cylinder.get('rental_days', 0) if isinstance(cylinder, dict) else 0)
            
            writer.writerow([
                display_id,
                safe_get(cylinder, 'serial_number'),
                safe_get(cylinder, 'type'),
                safe_get(cylinder, 'size'),
                safe_get(cylinder, 'status'),
                dispatch_date,
                rental_days,
                safe_get(cylinder, 'location'),
                safe_get(cylinder, 'pressure')
            ])
        
        # Summary statistics
        writer.writerow([])
        writer.writerow(['=== SUMMARY STATISTICS ==='])
        if customer_cylinders:
            rental_days_list = []
            for c in customer_cylinders:
                days = getattr(c, 'rental_days', 0) if hasattr(c, 'rental_days') else (c.get('rental_days', 0) if isinstance(c, dict) else 0)
                rental_days_list.append(days)
            
            total_days = sum(rental_days_list)
            avg_days = total_days // len(customer_cylinders) if customer_cylinders else 0
            longest_rental = max(rental_days_list) if rental_days_list else 0
            long_term_count = len([days for days in rental_days_list if days > 90])
            
            writer.writerow(['Total Cylinders:', len(customer_cylinders)])
            writer.writerow(['Average Days Dispatched:', avg_days])
            writer.writerow(['Longest Dispatch (Days):', longest_rental])
            writer.writerow(['Long-term Dispatches (90+ days):', long_term_count])
        else:
            writer.writerow(['No cylinders currently dispatched to this customer'])
        
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=customer_report_{safe_filename}_{timestamp}.csv'}
        )
    except Exception as e:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([f'Error generating customer report: {str(e)}'])
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename=customer_report_error_{timestamp}.csv'}
        )



def export_customer_pdf(customer, customer_cylinders, safe_filename, timestamp):
    """Export customer report as PDF"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        import io
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Build story
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        customer_name = customer.get('customer_name') or customer.get('name', 'Unknown Customer')
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=18, spaceAfter=30)
        story.append(Paragraph(f"Customer Report: {customer_name}", title_style))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Customer Details
        story.append(Paragraph("Customer Details", styles['Heading2']))
        customer_data = [
            ['Customer No:', customer.get('customer_no', '')],
            ['Name:', customer_name],
            ['Phone:', customer.get('customer_phone') or customer.get('phone', '')],
            ['Email:', customer.get('customer_email') or customer.get('email', '')],
            ['Address:', customer.get('customer_address') or customer.get('address', '')],
            ['City:', customer.get('customer_city', '')],
            ['State:', customer.get('customer_state', '')],
            ['Total Dispatched Cylinders:', str(len(customer_cylinders))]
        ]
        
        customer_table = Table(customer_data, colWidths=[2*inch, 4*inch])
        customer_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BACKGROUND', (1, 0), (1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(customer_table)
        story.append(Spacer(1, 20))
        
        # Dispatched Cylinders
        if customer_cylinders:
            story.append(Paragraph("Dispatched Cylinders (Sorted by Days Dispatched)", styles['Heading2']))
            
            cylinder_data = [['ID', 'Type', 'Size', 'Days Dispatched', 'Date Dispatched']]
            for cylinder in customer_cylinders:
                # Safe attribute access for SQLAlchemy objects
                def safe_get_cyl(obj, attr, default=''):
                    if hasattr(obj, attr):
                        return getattr(obj, attr) or default
                    return obj.get(attr, default) if isinstance(obj, dict) else default
                
                display_id = safe_get_cyl(cylinder, 'custom_id') or safe_get_cyl(cylinder, 'serial_number') or safe_get_cyl(cylinder, 'id')
                rental_days = getattr(cylinder, 'rental_days', 0) if hasattr(cylinder, 'rental_days') else (cylinder.get('rental_days', 0) if isinstance(cylinder, dict) else 0)
                dispatch_date = safe_get_cyl(cylinder, 'date_borrowed') or safe_get_cyl(cylinder, 'rental_date')
                if dispatch_date and len(str(dispatch_date)) >= 10:
                    dispatch_date = str(dispatch_date)[:10]
                
                cylinder_data.append([
                    str(display_id),
                    str(safe_get_cyl(cylinder, 'type')),
                    str(cylinder.get('size', '')),
                    str(cylinder.get('rental_days', 0)),
                    str(cylinder.get('date_borrowed', ''))
                ])
            
            cylinder_table = Table(cylinder_data, colWidths=[1*inch, 1.5*inch, 1.2*inch, 1.5*inch, 1.8*inch])
            cylinder_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(cylinder_table)
            story.append(Spacer(1, 20))
            
            # Summary Statistics
            story.append(Paragraph("Summary Statistics", styles['Heading2']))
            total_days = sum(c.get('rental_days', 0) for c in customer_cylinders)
            avg_days = total_days // len(customer_cylinders) if customer_cylinders else 0
            longest_rental = max(c.get('rental_days', 0) for c in customer_cylinders)
            long_term_count = len([c for c in customer_cylinders if c.get('rental_days', 0) > 90])
            
            summary_data = [
                ['Total Cylinders:', str(len(customer_cylinders))],
                ['Average Days Dispatched:', str(avg_days)],
                ['Longest Dispatch (Days):', str(longest_rental)],
                ['Long-term Dispatches (90+ days):', str(long_term_count)]
            ]
            
            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('BACKGROUND', (1, 0), (1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(summary_table)
        else:
            story.append(Paragraph("No cylinders currently dispatched to this customer", styles['Normal']))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        return Response(
            buffer.getvalue(),
            mimetype='application/pdf',
            headers={'Content-Disposition': f'attachment; filename=customer_report_{safe_filename}_{timestamp}.pdf'}
        )
    
    except ImportError:
        flash('PDF generation not available. Please use CSV format.', 'error')
        return redirect(url_for('reports'))

# Data Management Routes
@app.route('/admin/reset-data')
@login_required
def reset_data_page():
    """Show data reset confirmation page"""
    user_manager = UserManager()
    user = user_manager.get_user_by_id(session['user_id'])
    
    # Only admins can reset data
    if not user or user.get('role') != 'admin':
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('index'))
    
    # Get current data counts
    try:
        with CustomerService() as customer_service:
            customers, _ = customer_service.get_all(page=1, per_page=10000)
        with CylinderService() as cylinder_service:
            cylinders, _ = cylinder_service.get_all(page=1, per_page=10000)
        
        stats = {
            'total_customers': len(customers),
            'total_cylinders': len(cylinders),
            'active_rentals': len([c for c in cylinders if c.get('status', '').lower() == 'rented'])
        }
        
        return render_template('admin/reset_data.html', stats=stats)
    except Exception as e:
        flash(f'Error loading data statistics: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/admin/reset-data/confirm', methods=['POST'])
@login_required
def reset_data_confirm():
    """Reset all customer and cylinder data with backup"""
    user_manager = UserManager()
    user = user_manager.get_user_by_id(session['user_id'])
    
    # Only admins can reset data
    if not user or user.get('role') != 'admin':
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('index'))
    
    confirmation = request.form.get('confirmation')
    if confirmation != 'RESET ALL DATA':
        flash('Confirmation text does not match. Data not reset.', 'error')
        return redirect(url_for('reset_data_page'))
    
    try:
        # Create backup before reset
        backup_created = create_manual_backup('before_reset')
        
        if backup_created:
            try:
                # Reset customer data
                with CustomerService() as customer_service:
                    customers, _ = customer_service.get_all(page=1, per_page=10000)
                with CustomerService() as customer_service:
                    customers, _ = customer_service.get_all(page=1, per_page=10000)
                    pass  # TODO: Implement PostgreSQL data reset functionality
                
                # Reset cylinder data  
                with CylinderService() as cylinder_service:
                    # Clear all cylinders - this requires implementing a clear_all method or similar
                    pass  # TODO: Implement PostgreSQL data reset functionality
                
                # For now, show message that reset is not implemented for PostgreSQL
                flash('PostgreSQL data reset functionality needs to be implemented. Please use SQL tools to reset data if needed.', 'warning')
            except Exception as e:
                flash(f'Error during data reset: {str(e)}', 'error')
        else:
            flash('Failed to create backup. Data reset cancelled for safety.', 'error')
            
    except Exception as e:
        flash(f'Error during data reset: {str(e)}', 'error')
    
    return redirect(url_for('index'))

@app.route('/admin/backup-data')
@login_required
def manual_backup():
    """Create manual backup of all data"""
    user_manager = UserManager()
    user = user_manager.get_user_by_id(session['user_id'])
    
    # Only admins can create backups
    if not user or user.get('role') != 'admin':
        flash('Access denied. Admin privileges required.', 'error')
        return redirect(url_for('index'))
    
    try:
        backup_created = create_manual_backup('manual')
        if backup_created:
            flash('Manual backup created successfully.', 'success')
        else:
            flash('Failed to create backup.', 'error')
    except Exception as e:
        flash(f'Error creating backup: {str(e)}', 'error')
    
    return redirect(url_for('index'))

def create_manual_backup(backup_type='manual'):
    """Create backup of all data files"""
    try:
        # Create backups directory if it doesn't exist
        backup_dir = 'backups'
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)
        
        # Create timestamped backup directory
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_subdir = os.path.join(backup_dir, f'{backup_type}_backup_{timestamp}')
        os.makedirs(backup_subdir)
        
        # Backup data files
        data_files = ['customers.json', 'cylinders.json', 'users.json']
        for file in data_files:
            src_path = os.path.join('data', file)
            if os.path.exists(src_path):
                dst_path = os.path.join(backup_subdir, file)
                shutil.copy2(src_path, dst_path)
        
        # Create backup info file
        info_file = os.path.join(backup_subdir, 'backup_info.json')
        backup_info = {
            'backup_type': backup_type,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'files_backed_up': data_files,
            'system': 'Varasai Oxygen'
        }
        with open(info_file, 'w') as f:
            json.dump(backup_info, f, indent=2)
        
        return True
    except Exception as e:
        print(f"Backup creation failed: {str(e)}")
        return False

# Auto-backup system
class AutoBackupManager:
    """Manages automatic backup system"""
    
    def __init__(self):
        self.backup_interval = 14 * 24 * 60 * 60  # 2 weeks in seconds
        self.last_backup_file = 'data/last_backup.json'
        self.running = False
        self.thread = None
    
    def start_auto_backup(self):
        """Start the automatic backup system"""
        if not self.running:
            self.running = True
            self.thread = threading.Thread(target=self._backup_loop, daemon=True)
            self.thread.start()
    
    def stop_auto_backup(self):
        """Stop the automatic backup system"""
        self.running = False
    
    def _backup_loop(self):
        """Main backup loop running in background"""
        while self.running:
            try:
                if self._should_create_backup():
                    self._create_auto_backup()
                # Check every hour
                time.sleep(3600)
            except Exception as e:
                print(f"Auto-backup error: {str(e)}")
                time.sleep(3600)  # Wait an hour before trying again
    
    def _should_create_backup(self):
        """Check if backup should be created"""
        try:
            if not os.path.exists(self.last_backup_file):
                return True
            
            with open(self.last_backup_file, 'r') as f:
                last_backup_info = json.load(f)
            
            last_backup_time = datetime.strptime(
                last_backup_info['last_backup'], 
                '%Y-%m-%d %H:%M:%S'
            )
            
            time_since_backup = datetime.now() - last_backup_time
            return time_since_backup.total_seconds() >= self.backup_interval
            
        except Exception:
            return True  # If can't read file, assume backup needed
    
    def _create_auto_backup(self):
        """Create automatic backup"""
        try:
            backup_created = create_manual_backup('auto')
            if backup_created:
                # Update last backup time
                backup_info = {
                    'last_backup': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    'backup_type': 'automatic',
                    'next_backup': (datetime.now() + timedelta(days=14)).strftime('%Y-%m-%d %H:%M:%S')
                }
                
                # Ensure data directory exists
                if not os.path.exists('data'):
                    os.makedirs('data')
                
                with open(self.last_backup_file, 'w') as f:
                    json.dump(backup_info, f, indent=2)
                
                print(f"Auto-backup created successfully at {datetime.now()}")
        except Exception as e:
            print(f"Auto-backup failed: {str(e)}")

# Initialize auto-backup system
auto_backup_manager = AutoBackupManager()

# Start auto-backup when app starts (using app context)
def initialize_auto_backup():
    """Initialize automatic backup system"""
    auto_backup_manager.start_auto_backup()

# Initialize auto-backup at import time
with app.app_context():
    initialize_auto_backup()

# PDF Export Routes - Temporarily disabled due to syntax errors
@app.route('/export/customers.pdf')
@login_required
def export_customers_pdf():
    """Export all customers to PDF"""
    flash('PDF generation temporarily disabled. Please use CSV format.', 'warning')
    return redirect(url_for('reports'))

@app.route('/export/cylinders.pdf')
@login_required
def export_cylinders_pdf():
    """Export all cylinders to PDF"""
    flash('PDF generation temporarily disabled. Please use CSV format.', 'warning')
    return redirect(url_for('reports'))
    
    # Title
    title = Paragraph("Varasai Oxygen - Cylinder Inventory Report", styles['Title'])
    story.append(title)
    story.append(Spacer(1, 12))
    
    # Date and summary
    date_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    date_para = Paragraph(f"Generated on: {date_str}", styles['Normal'])
    story.append(date_para)
    
    summary_para = Paragraph(f"Total Cylinders: {len(cylinders)}", styles['Normal'])
    story.append(summary_para)
    
    # Status breakdown
    status_counts = {}
    for cylinder in cylinders:
        status = cylinder.get('status', 'Unknown')
        status_counts[status] = status_counts.get(status, 0) + 1
    
    status_text = " | ".join([f"{status}: {count}" for status, count in status_counts.items()])
    status_para = Paragraph(f"Status Breakdown: {status_text}", styles['Normal'])
    story.append(status_para)
    story.append(Spacer(1, 12))
    
    # Cylinder table
    if cylinders:
        data = [['ID', 'Type', 'Size', 'Status', 'Location', 'Customer']]
        for cylinder in cylinders:
            # Use custom ID if available, otherwise fallback to generated serial
            display_id = cylinder_model.get_display_id(cylinder)
            
            data.append([
                display_id[:15],
                cylinder.get('type', '')[:15],
                cylinder.get('size', '')[:12],
                cylinder.get('status', '')[:10],
                cylinder.get('location', '')[:15],
                cylinder.get('customer_name', '')[:15]
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(table)
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    return Response(
        buffer.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename=cylinders_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'}
    )

@app.route('/export/rental-activities.pdf')
@login_required
def export_rental_activities_pdf():
    """Export rental activities to PDF"""
    flash('PDF generation temporarily disabled. Please use CSV format.', 'warning')
    return redirect(url_for('reports'))


@app.route('/export/rental_history')
@login_required
def export_rental_history():
    """Export complete rental history to Excel"""
    import io
    from openpyxl import Workbook
    from db_service import RentalHistoryService
    
    try:
        with RentalHistoryService() as service:
            all_history, _ = service.get_all()  # Get all history
        
        workbook = Workbook()
        
        # Create separate sheets for active and past rentals
        active_sheet = workbook.active
        active_sheet.title = "Active Rentals"
        past_sheet = workbook.create_sheet("Past Rentals")
        
        # Headers
        headers = [
            'Customer No', 'Customer Name', 'Customer Phone', 'Customer Address',
            'Cylinder ID', 'Cylinder Type', 'Cylinder Size', 
            'Dispatch Date', 'Return Date', 'Rental Days'
        ]
        
        # Active rentals sheet
        for col, header in enumerate(headers, 1):
            active_sheet.cell(row=1, column=col, value=header)
        
        row = 2
        for record in all_history['active']:
            active_sheet.cell(row=row, column=1, value=record.get('customer_no', ''))
            active_sheet.cell(row=row, column=2, value=record.get('customer_name', ''))
            active_sheet.cell(row=row, column=3, value=record.get('customer_phone', ''))
            active_sheet.cell(row=row, column=4, value=record.get('customer_address', ''))
            active_sheet.cell(row=row, column=5, value=record.get('cylinder_custom_id', '') or record.get('cylinder_serial', ''))
            active_sheet.cell(row=row, column=6, value=record.get('cylinder_type', ''))
            active_sheet.cell(row=row, column=7, value=record.get('cylinder_size', ''))
            active_sheet.cell(row=row, column=8, value=record.get('date_borrowed', '')[:10] if record.get('date_borrowed') else '')
            active_sheet.cell(row=row, column=9, value='')  # No return date for active
            active_sheet.cell(row=row, column=10, value=record.get('rental_days', 0))
            row += 1
        
        # Past rentals sheet
        for col, header in enumerate(headers, 1):
            past_sheet.cell(row=1, column=col, value=header)
        
        row = 2
        for record in all_history['past']:
            past_sheet.cell(row=row, column=1, value=record.get('customer_no', ''))
            past_sheet.cell(row=row, column=2, value=record.get('customer_name', ''))
            past_sheet.cell(row=row, column=3, value=record.get('customer_phone', ''))
            past_sheet.cell(row=row, column=4, value=record.get('customer_address', ''))
            past_sheet.cell(row=row, column=5, value=record.get('cylinder_custom_id', '') or record.get('cylinder_serial', ''))
            past_sheet.cell(row=row, column=6, value=record.get('cylinder_type', ''))
            past_sheet.cell(row=row, column=7, value=record.get('cylinder_size', ''))
            past_sheet.cell(row=row, column=8, value=record.get('date_borrowed', '')[:10] if record.get('date_borrowed') else '')
            past_sheet.cell(row=row, column=9, value=record.get('date_returned', '')[:10] if record.get('date_returned') else '')
            past_sheet.cell(row=row, column=10, value=record.get('rental_days', 0))
            row += 1
        
        # Save to memory
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        filename = f"rental_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error exporting rental history: {str(e)}', 'error')
        return redirect(url_for('reports'))

@app.route('/wipe_records', methods=['GET', 'POST'])
@login_required  
@admin_required
def wipe_records():
    """Wipe all records with backup option"""
    if request.method == 'POST':
        wipe_option = request.form.get('wipe_option')
        create_backup = request.form.get('create_backup') == 'on'
        
        try:
            if create_backup:
                # Create backup before wiping
                backup_filename = f"backup_before_wipe_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                backup_data = {}
                
                # Export all data
                with CustomerService() as customer_service:
                    customers, _ = customer_service.get_all(page=1, per_page=100000)
                    backup_data['customers'] = []
                    for c in customers:
                        customer_dict = {
                            'id': c.id if hasattr(c, 'id') else c.get('id'),
                            'customer_no': c.customer_no if hasattr(c, 'customer_no') else c.get('customer_no'),
                            'customer_name': c.customer_name if hasattr(c, 'customer_name') else c.get('customer_name'),
                            'customer_email': c.customer_email if hasattr(c, 'customer_email') else c.get('customer_email'),
                            'customer_phone': c.customer_phone if hasattr(c, 'customer_phone') else c.get('customer_phone'),
                            'customer_address': c.customer_address if hasattr(c, 'customer_address') else c.get('customer_address'),
                            'customer_city': c.customer_city if hasattr(c, 'customer_city') else c.get('customer_city'),
                            'customer_state': c.customer_state if hasattr(c, 'customer_state') else c.get('customer_state')
                        }
                        backup_data['customers'].append(customer_dict)
                
                with CylinderService() as cylinder_service:
                    cylinders, _ = cylinder_service.get_all(page=1, per_page=100000)
                    backup_data['cylinders'] = []
                    for c in cylinders:
                        cylinder_dict = {
                            'id': c.id if hasattr(c, 'id') else c.get('id'),
                            'custom_id': c.custom_id if hasattr(c, 'custom_id') else c.get('custom_id'),
                            'serial_number': c.serial_number if hasattr(c, 'serial_number') else c.get('serial_number'),
                            'type': c.type if hasattr(c, 'type') else c.get('type'),
                            'size': c.size if hasattr(c, 'size') else c.get('size'),
                            'status': c.status if hasattr(c, 'status') else c.get('status'),
                            'location': c.location if hasattr(c, 'location') else c.get('location'),
                            'rented_to': c.rented_to if hasattr(c, 'rented_to') else c.get('rented_to'),
                            'customer_name': c.customer_name if hasattr(c, 'customer_name') else c.get('customer_name'),
                            'customer_email': c.customer_email if hasattr(c, 'customer_email') else c.get('customer_email'),
                            'customer_phone': c.customer_phone if hasattr(c, 'customer_phone') else c.get('customer_phone'),
                            'date_borrowed': str(c.date_borrowed) if hasattr(c, 'date_borrowed') and c.date_borrowed else (str(c.get('date_borrowed')) if c.get('date_borrowed') else None),
                            'rental_date': str(c.rental_date) if hasattr(c, 'rental_date') and c.rental_date else (str(c.get('rental_date')) if c.get('rental_date') else None),
                            'date_returned': str(c.date_returned) if hasattr(c, 'date_returned') and c.date_returned else (str(c.get('date_returned')) if c.get('date_returned') else None)
                        }
                        backup_data['cylinders'].append(cylinder_dict)
                
                # Save backup file
                backup_path = f"data/{backup_filename}"
                os.makedirs('data', exist_ok=True)
                with open(backup_path, 'w') as f:
                    json.dump(backup_data, f, indent=2, default=str)
                
                flash(f'Backup created: {backup_filename}', 'info')
            
            # Perform the wipe
            with app.app_context():
                from app import db
                from sqlalchemy import text
                
                if wipe_option == 'all':
                    # Wipe all records
                    db.session.execute(text("DELETE FROM rental_history"))
                    db.session.execute(text("DELETE FROM cylinders"))  
                    db.session.execute(text("DELETE FROM customers"))
                    flash('All records have been wiped', 'success')
                    
                elif wipe_option == 'cylinders':
                    # Wipe only cylinders and rental history
                    db.session.execute(text("DELETE FROM rental_history"))
                    db.session.execute(text("DELETE FROM cylinders"))
                    flash('All cylinder records have been wiped', 'success')
                    
                elif wipe_option == 'customers':
                    # Wipe only customers and update cylinders
                    db.session.execute(text("UPDATE cylinders SET rented_to = NULL WHERE rented_to IS NOT NULL"))
                    db.session.execute(text("DELETE FROM rental_history"))
                    db.session.execute(text("DELETE FROM customers"))
                    flash('All customer records have been wiped, cylinders updated', 'success')
                
                db.session.commit()
                
        except Exception as e:
            flash(f'Error during wipe operation: {str(e)}', 'error')
            return redirect(url_for('wipe_records'))
            
        return redirect(url_for('index'))
    
    # GET request - show confirmation form
    return render_template('wipe_records.html')

@app.route('/direct_db_import', methods=['GET', 'POST'])
@login_required
@admin_required
def direct_db_import():
    """Import data directly from external PostgreSQL database"""
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'connect':
            connection_string = request.form.get('connection_string')
            session['db_connection'] = connection_string
            
            # Test connection
            from direct_db_import import DirectDatabaseImporter
            with DirectDatabaseImporter() as importer:
                if importer.connect_to_external_db(connection_string):
                    tables = importer.list_tables()
                    session['available_tables'] = tables
                    flash(f'Connected successfully! Found {len(tables)} tables.', 'success')
                    return render_template('direct_db_import.html', 
                                         connected=True, 
                                         tables=tables)
                else:
                    flash('Failed to connect to database. Check connection string.', 'error')
                    return render_template('direct_db_import.html')
        
        elif action == 'preview':
            table_name = request.form.get('table_name')
            connection_string = session.get('db_connection')
            
            if not connection_string:
                flash('No database connection. Please connect first.', 'error')
                return redirect(url_for('direct_db_import'))
            
            from direct_db_import import DirectDatabaseImporter
            with DirectDatabaseImporter() as importer:
                if importer.connect_to_external_db(connection_string):
                    preview_data = importer.preview_table(table_name)
                    return render_template('direct_db_import.html',
                                         connected=True,
                                         tables=session.get('available_tables', []),
                                         preview_table=table_name,
                                         preview_data=preview_data)
        
        elif action == 'import':
            table_name = request.form.get('table_name')
            import_type = request.form.get('import_type')
            connection_string = session.get('db_connection')
            
            # Get field mappings from form
            field_mapping = {}
            for key in request.form.keys():
                if key.startswith('mapping_'):
                    source_field = key.replace('mapping_', '')
                    target_field = request.form.get(key)
                    if target_field and target_field != 'skip':
                        field_mapping[source_field] = target_field
            
            if not field_mapping:
                flash('Please map at least one field.', 'error')
                return redirect(url_for('direct_db_import'))
            
            from direct_db_import import DirectDatabaseImporter
            with DirectDatabaseImporter() as importer:
                if importer.connect_to_external_db(connection_string):
                    if import_type == 'customers':
                        result = importer.import_customers_from_table(table_name, field_mapping)
                    else:
                        result = importer.import_cylinders_from_table(table_name, field_mapping)
                    
                    if result['success']:
                        flash(f'Import successful! Imported: {result["imported"]}, Updated: {result["updated"]}', 'success')
                    else:
                        flash(f'Import completed with errors: {result.get("error", "Unknown error")}', 'warning')
                        if result.get('errors'):
                            for error in result['errors'][:5]:  # Show first 5 errors
                                flash(f'Error: {error}', 'info')
                    
                    return redirect(url_for('direct_db_import'))
    
    # GET request
    return render_template('direct_db_import.html')

@app.route('/migrate_to_render', methods=['GET', 'POST'])
@login_required
@admin_required
def migrate_to_render():
    """Export data for Render migration"""
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'export_all':
            from migrate_to_render import ReplitToRenderMigrator
            
            migrator = ReplitToRenderMigrator()
            result = migrator.create_render_migration_package()
            
            if result.get('success'):
                flash(f'Migration package created successfully in {result["migration_dir"]}/', 'success')
                flash(f'Exported {result["stats"]["total_customers"]} customers and {result["stats"]["total_cylinders"]} cylinders', 'info')
                return render_template('migrate_to_render.html', 
                                     export_result=result,
                                     show_download=True)
            else:
                flash(f'Export failed: {result.get("error")}', 'error')
                return render_template('migrate_to_render.html')
        
        elif action == 'get_connection_info':
            from migrate_to_render import ReplitToRenderMigrator
            
            migrator = ReplitToRenderMigrator()
            connection_info = migrator.get_render_connection_info()
            
            return render_template('migrate_to_render.html',
                                 connection_info=connection_info,
                                 show_connection_info=True)
    
    # GET request
    return render_template('migrate_to_render.html')
